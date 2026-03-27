"""Excel report generation service – produces detailed professional .xlsx reports.

Features:
- Dashboard sheet with KPI boxes & multiple charts
- Professional formatting with branded colors & conditional formatting
- Multiple worksheets: Dashboard, Summary, Price Data, Technical Indicators,
  Risk Analysis, Model Details, Backtest, Walk-Forward, Error Analysis,
  Statistical Tests, Conclusion
- Embedded charts (BarChart, LineChart, DoughnutChart, AreaChart, etc.)
- Technical indicators (SMA, EMA, RSI, MACD, Bollinger Bands)
- Risk metrics (VaR, max drawdown, Sharpe-like ratio)
- Multi-language support (English, Slovak)
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.logging_config import get_logger
from app.schemas.schemas import GenerateReportRequest, ReportItem
from app.services.ohlc_service import get_close_series, get_ohlc_dataframe

log = get_logger(__name__)

# ─── Colors ────────────────────────────────────────────────────────────────────

BRAND_BLUE = "0EA5E9"
BRAND_DARK = "1E293B"
HEADER_BG = "0F172A"
HEADER_FG = "FFFFFF"
GOOD_GREEN = "16A34A"
WARN_AMBER = "D97706"
BAD_RED = "DC2626"
LIGHT_BG = "F1F5F9"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "DCFCE7"
LIGHT_RED = "FEE2E2"
LIGHT_AMBER = "FEF3C7"
BORDER_COLOR = "CBD5E1"
KPI_BG_1 = "EFF6FF"
KPI_BG_2 = "F0FDF4"
KPI_BG_3 = "FFF7ED"
KPI_BG_4 = "FAF5FF"
KPI_BG_5 = "F0F9FF"
ACCENT_PURPLE = "7C3AED"
ACCENT_TEAL = "0D9488"

# ─── Translations ──────────────────────────────────────────────────────────────

TRANSLATIONS = {
    "en": {
        "title": "FX Analytics Report",
        "dashboard_sheet": "Dashboard",
        "summary_sheet": "Summary",
        "data_sheet": "Price Data",
        "indicators_sheet": "Technical Indicators",
        "risk_sheet": "Risk Analysis",
        "models_sheet": "Model Details",
        "backtest_sheet": "Backtest Results",
        "windows_sheet": "Walk-Forward Windows",
        "error_sheet": "Error Analysis",
        "tests_sheet": "Statistical Tests",
        "conclusion_sheet": "Conclusion",
        "overview": "Analysis Overview",
        "pair": "Currency Pair",
        "timeframe": "Timeframe",
        "period": "Analysis Period",
        "data_points": "Data Points",
        "generated": "Generated",
        "mean": "Mean Close",
        "std": "Std Dev",
        "min": "Minimum",
        "max": "Maximum",
        "range": "Price Range",
        "model_col": "Model",
        "mae_col": "MAE",
        "rmse_col": "RMSE",
        "da_col": "Directional Accuracy",
        "rank_col": "Rank",
        "best_model": "Best Model",
        "metric_desc": "Metric Description",
        "mae_desc": "Mean Absolute Error – average magnitude of prediction errors",
        "rmse_desc": "Root Mean Squared Error – penalizes large errors more",
        "da_desc": "Percentage of correctly predicted price direction (up/down)",
        "naive_desc": "Baseline: predicts next value = last observed",
        "ma_desc": "Average of last N closing prices (window = 20)",
        "arima_desc": "ARIMA – parametric time-series model capturing trends and autocorrelation",
        "ridge_desc": "Regularized linear regression using technical indicators as features",
        "rf_desc": "Ensemble of decision trees that captures non-linear patterns",
        "ai_desc": "AI-enhanced ensemble combining base model predictions with GPT analysis (author's contribution)",
        "train_start": "Train Start",
        "train_end": "Train End",
        "test_start": "Test Start",
        "test_end": "Test End",
        "window_num": "Window #",
        "adf_title": "ADF Test (Stationarity)",
        "adf_desc": "Tests whether the time series has a unit root. A low p-value (<0.05) means the series is stationary.",
        "lb_title": "Ljung-Box Test (Autocorrelation)",
        "lb_desc": "Tests whether residuals exhibit autocorrelation. A high p-value (>0.05) means no significant autocorrelation.",
        "dm_title": "Diebold-Mariano Test",
        "dm_desc": "Compares predictive accuracy of two competing models.",
        "statistic": "Test Statistic",
        "pvalue": "p-value",
        "result": "Result",
        "stationary": "Stationary",
        "not_stationary": "Non-stationary",
        "no_autocorr": "No significant autocorrelation",
        "autocorr_detected": "Autocorrelation detected",
        "better_model": "Better Model",
        "significant": "Significant (p < 0.05)",
        "not_significant": "Not significant (p >= 0.05)",
        "explanation": "Explanation",
        "conclusion": "Conclusion",
        "comparison_notes": "Comparison Notes",
        "note_da_50": "Directional accuracy ≈50% is no better than random guessing",
        "note_da_55": "Directional accuracy >55% indicates meaningful predictive ability",
        "note_rmse_best": "Lower RMSE = better at minimizing prediction error magnitude",
        "note_da_vs_naive": "A model should beat Naive's DA to be considered genuinely better",
        # ── Keys matching Slovak additions (for English fallback) ──
        "price_statistics": "Price Statistics",
        "quantile_analysis": "Quantile Analysis",
        "return_statistics": "Return Statistics",
        "monthly_performance": "Monthly Performance",
        "positive_months": "Positive Months",
        "avg_monthly_return": "Average Monthly Return",
        "best_month": "Best Month",
        "worst_month": "Worst Month",
        "month_col": "Month",
        "return_pct": "Return %",
        "direction": "Direction",
        "median_close": "Median Close",
        "first_close": "First Close",
        "last_close": "Last Close",
        "total_change_pct": "Total Change %",
        "mean_log_return": "Mean Log Return",
        "std_log_return": "Std Log Return",
        "daily_volatility": "Daily Volatility",
        "annualized_volatility": "Annualized Volatility",
        "skewness": "Skewness",
        "kurtosis_excess": "Kurtosis (excess)",
        "min_return": "Min Return",
        "max_return": "Max Return",
        "positive_days": "Positive Days",
        "positive_days_pct": "Positive Days %",
        "largest_up": "Largest Up Move",
        "largest_down": "Largest Down Move",
        "jb_statistic": "Jarque-Bera Statistic",
        "jb_pvalue": "Jarque-Bera p-value",
        "distribution_label": "Distribution",
        "dist_normal": "Normal",
        "dist_non_normal": "Non-normal (heavy tails)",
        "sw_statistic": "Shapiro-Wilk Statistic",
        "sw_pvalue": "Shapiro-Wilk p-value",
        "volatility_analysis": "Volatility Analysis",
        "daily_vol_sigma": "Daily Volatility (sigma)",
        "ann_vol_sigma": "Annualized Volatility (sigma*sqrt(252))",
        "intraday_range_avg": "Intraday Range (avg)",
        "vol_regime": "Volatility Regime",
        "vol_low": "Low Volatility",
        "vol_normal": "Normal Volatility",
        "vol_high": "High Volatility",
        "var_title": "Value at Risk (VaR)",
        "confidence": "Confidence",
        "var_daily": "VaR (Daily)",
        "var_annualized": "VaR (Annualized)",
        "expected_shortfall": "Expected Shortfall",
        "drawdown_analysis": "Drawdown Analysis",
        "max_drawdown": "Max Drawdown",
        "max_dd_duration": "Max Drawdown Duration (bars)",
        "current_drawdown": "Current Drawdown",
        "avg_drawdown": "Average Drawdown",
        "models_used": "Models Used in Analysis",
        "model_type": "Type",
        "complexity_col": "Complexity",
        "metrics_explained": "Metrics Explained",
        "comparison_vs_naive": "Detailed Model vs Naive Comparison",
        "rmse_vs_naive": "RMSE vs Naive",
        "mae_vs_naive": "MAE vs Naive",
        "da_vs_naive": "DA vs Naive",
        "verdict": "Verdict",
        "outperforms": "Outperforms",
        "partial_result": "Partial",
        "underperforms": "Underperforms",
        "per_model_stats": "Per-Model Window Statistics",
        "error_dist_per_model": "Error Distribution per Model",
        "mean_error": "Mean Error",
        "std_error": "Std Error",
        "min_error": "Min Error",
        "max_error": "Max Error",
        "median_error": "Median Error",
        "error_skewness": "Error Skewness",
        "error_kurtosis": "Error Kurtosis",
        "mae_percentiles": "MAE Percentiles per Model",
        "consistency_score": "Consistency Score",
        "consistency_desc": "A lower coefficient of variation (CV) means more consistent predictions across windows.",
        "cons_excellent": "Excellent",
        "cons_good": "Good",
        "cons_moderate": "Moderate",
        "cons_poor": "Poor",
        "significance_summary": "Statistical Significance Summary",
        "key_findings": "Key Findings",
        "final_ranking": "Final Model Ranking",
        "methodology_title": "Methodology",
        "beats_naive_col": "Beats Naive?",
        "baseline_label": "baseline",
        "yes_label": "Yes",
        "no_label": "No",
        "best_label": "BEST",
        "strong_label": "Strong",
        "moderate_label": "Moderate",
        "weak_label": "Weak",
        "target_info_title": "Prediction Target",
        "target_close": "Predicting closing price (Close) in currency units",
        "target_reconstr": "Price reconstruction: close(t+1) = close(t) * exp(r(t+1))",
        "error_definition": "Error Definition",
        "error_on_close": "Error (MAE, RMSE) measured in price units (e.g. USD)",
        "data_source_title": "Data Source",
        "data_provider": "Provider",
        "data_api": "API / Library",
        "data_timezone": "Timezone",
        "data_cleaning": "Data Cleaning",
        "data_cleaning_desc": "Automatic detection of duplicates, invalid bars (OHLC=0/NaN), monotonicity check. All timestamps in UTC.",
        "da_formula_title": "Directional Accuracy (DA) Calculation",
        "da_formula": "DA = (1/n) * sum( 1[sign(r_hat_t) = sign(r_t)] )",
        "da_baseline_text": "Random level: ~50% (equivalent to coin flip)",
        "da_interp": "DA > 55% indicates meaningful predictive ability; DA < 50% is worse than random.",
        "model_params": "Model Parameters",
        "features_used": "Features Used",
        "train_val_split": "Train/Validation Split",
        "determinism": "Determinism",
        "ai_composition": "AI Ensemble Composition",
        "ai_how": "Prediction Combination Method",
        "ai_comp_detail": "Predictions from Naive, MovingAverage, ARIMA, Ridge, RandomForest + technical indicators (RSI, MACD, Momentum, Bollinger) + last 30 prices.",
        "ai_how_detail": "LLM (GPT-4o-mini) analyzes context and returns weighted prediction. Fallback: simple average if API unavailable. temperature=0.",
        "selection_rule_title": "Best Model Selection Rule",
        "selection_rule_desc": "From models whose DA >= DA(Naive), select the one with lowest RMSE. If none meet this condition, select the model with highest DA.",
        "stat_tests_conclusion": "Statistical Tests Conclusions",
        "limitations_title": "Limitations",
        "improvements_title": "Suggestions for Improvement",
        "top_errors_title": "Top Errors (by window MAE)",
        "window_label": "Window",
        "notes_col": "Notes",
        "lb_residuals": "Model Residuals",
        "lb_lag": "Number of Lags",
        "rolling_perf": "Rolling Performance",
        # ── Model types & complexity ──
        "type_baseline": "Baseline",
        "type_statistical": "Statistical",
        "type_timeseries": "Time series",
        "type_ml": "Machine learning",
        "type_ai_ensemble": "AI Ensemble",
        "complexity_low": "Low",
        "complexity_medium": "Medium",
        "complexity_high": "High",
        "complexity_very_high": "Very High",
        # ── Model parameters detail ──
        "params_none": "No parameters",
        "features_naive": "None — uses only last price",
        "deterministic_yes": "Yes (deterministic)",
        "params_ma_window": "window k = 20",
        "features_ma": "Last 20 closing prices",
        "params_arima": "ARIMA(p,d,q) — automatic selection via AIC (pmdarima), typically (1,1,1)",
        "features_arima": "Close price series (univariate model)",
        "deterministic_arima": "Yes (fixed seed, same data = same result)",
        "train_val_desc": "Time-based split: last 20% of data = validation",
        "deterministic_ridge": "Yes (deterministic, no randomness)",
        "deterministic_rf": "Yes (random_state=42)",
        "deterministic_ai": "Partial (temperature=0, but LLM may vary)",
        "formula_label": "Formula",
        "baseline_da_label": "Baseline",
        "interpretation_label": "Interpretation",
        # ── Backtest notes ──
        "vs_naive_da": "vs Naive DA",
        "rmse_rank_col": "RMSE Rank",
        "note_lowest_rmse": "Lowest RMSE",
        "note_highest_da": "Highest DA",
        "note_below_random": "Below random level",
        "note_beats_naive": "Beats Naive in both metrics",
        "chart_mae_rmse": "MAE & RMSE Comparison",
        "chart_da_pct": "Directional Accuracy (%)",
        # ── Statistical significance ──
        "lb_residuals_of": "Residuals of model: {}",
        "lb_lag_default": "10 (default)",
        "sig_adf_ok": "Stationary — log returns are suitable for modeling.",
        "sig_adf_fail": "Non-stationary — caution, series may have a trend.",
        "sig_lb_ok": "No autocorrelation — residuals behave as white noise.",
        "sig_lb_fail": "Autocorrelation detected — model may be insufficient.",
        "sig_dm_ok": "Significant — {} is statistically better.",
        "sig_dm_fail": "Not significant — no statistical difference between models.",
        # ── Conclusion findings ──
        "finding_best_model": "Best overall model: {} (RMSE: {}, DA: {}%)",
        "finding_model_count": "Number of tested models: {}",
        "finding_data_points": "Data points analyzed: {}",
        "finding_period": "Period: {} — {}",
        "finding_naive_da": "Baseline DA (Naive): {}%",
        "finding_ann_vol": "Annualized volatility: {}%",
        "finding_max_dd": "Maximum drawdown: {}%",
        "finding_total_change": "Total price change: {}%",
        "finding_winners": "Models outperforming Naive in both metrics: {}",
        "finding_no_winners": "No model outperforms Naive in both RMSE and DA.",
        "finding_da_below_50": "Warning: Best model DA ({}%) is below 50% — worse than random.",
        "finding_da_near_50": "Best model DA ({}%) is near random level — limited predictive power.",
        # ── Conclusion: stat tests ──
        "conc_adf_stationary": "stationary",
        "conc_adf_nonstationary": "non-stationary",
        "conc_adf_ok_comment": "Suitable for modeling.",
        "conc_adf_fail_comment": "Caution: further differencing may be needed.",
        "conc_lb_ok": "no significant autocorrelation in residuals",
        "conc_lb_fail": "autocorrelation detected — model may be insufficient",
        "conc_dm_ok": "significant difference, better model: {}",
        "conc_dm_fail": "no statistically significant difference between models",
        # ── Methodology ──
        "method_walkforward": "Walk-forward cross-validation with rolling windows",
        "method_windows": "Training window: {} bars, Test window: {} bars",
        "method_metrics": "Metrics: MAE, RMSE, Directional Accuracy (DA)",
        "method_tests": "Statistical tests: ADF (stationarity), Ljung-Box (autocorrelation), Diebold-Mariano (model comparison)",
        "method_datasource": "Data source: Yahoo Finance (yfinance)",
        "method_generated": "Report generated: {}",
        # ── Limitations ──
        "lim_timeframe": "Only {} timeframe analyzed for {}",
        "lim_interval": "Limited time interval: {} to {} ({} observations)",
        "lim_no_exogenous": "No exogenous variables (interest rates, macroeconomic indicators, sentiment)",
        "lim_yahoo": "Yahoo Finance is not a primary data source (aggregates from various providers)",
        "lim_ai_api": "AI ensemble depends on external API (may vary)",
        # ── Improvements ──
        "imp_more_pairs": "Extend to more currency pairs (GBP/USD, USD/JPY) and longer periods",
        "imp_more_features": "Add more features: volatility clustering (GARCH), sentiment from financial news",
        "imp_deep_models": "Implement deep models (LSTM, GRU, Transformer) for comparison",
        "imp_validation": "More robust validation: expanding window, multiple walk-forward",
        "imp_exogenous": "Add exogenous variables (interest rates, CPI, PMI)",
        "imp_oos_verify": "Verify result consistency on out-of-sample data from another source (ECB, FRED)",
        # ── Indicator title ──
        "current_indicator_values": "Current indicator values (last bar)",
    },
    "sk": {
        "title": "FX Analytics Report",
        "dashboard_sheet": "Dashboard",
        "summary_sheet": "Suhrn",
        "data_sheet": "Cenove udaje",
        "indicators_sheet": "Technicke indikatory",
        "risk_sheet": "Analyza rizika",
        "models_sheet": "Detaily modelov",
        "backtest_sheet": "Vysledky backtestu",
        "windows_sheet": "Walk-Forward okna",
        "error_sheet": "Analyza chyb",
        "tests_sheet": "Statisticke testy",
        "conclusion_sheet": "Zaver",
        "overview": "Prehlad analyzy",
        "pair": "Menovy par",
        "timeframe": "Casovy ramec",
        "period": "Obdobie analyzy",
        "data_points": "Pocet pozorovaní",
        "generated": "Vygenerovane",
        "mean": "Priemer Close",
        "std": "Smerodajna odchylka",
        "min": "Minimum",
        "max": "Maximum",
        "range": "Cenovy rozsah",
        "model_col": "Model",
        "mae_col": "MAE",
        "rmse_col": "RMSE",
        "da_col": "Smerova presnost",
        "rank_col": "Poradie",
        "best_model": "Najlepsi model",
        "metric_desc": "Popis metriky",
        "mae_desc": "Stredna absolutna chyba - priemerna velkost chyb predikcie",
        "rmse_desc": "Odmocnina strednej stvorcovej chyby - viac penalizuje velke chyby",
        "da_desc": "Percento spravne predpovedaneho smeru ceny (hore/dole). DA = (1/n) * sum(1[sign(r_hat) = sign(r)]). Nahodna uroven ~ 50%.",
        "naive_desc": "Zaklad (nahodna prechadzka): predpoved dalsej hodnoty = posledna pozorovana. P(t+1) = P(t).",
        "ma_desc": "Priemer poslednych N uzatvaracich cien (okno = 20). P(t+1) = (1/k) * sum(P(t-i)).",
        "arima_desc": "ARIMA(1,1,1) - parametricky model casoveho radu. Automaticky vyber p,d,q cez AIC (pmdarima). Zachytava trendy a autokorelaciu.",
        "ridge_desc": "Ridge regresia (L2 regularizacia, alpha=1.0). Vstupy: 5 lag returns, SMA(20), EMA(20), RSI(14), MACD(12,26,9). Pipeline: StandardScaler + Ridge.",
        "rf_desc": "Random Forest (100 stromov, max_depth=10). Vstupy: 5 lag returns, SMA(20), EMA(20), RSI(14), MACD(12,26,9). Rekurzivna predikcia.",
        "ai_desc": "AI ansambel - hlavny prinos autora. Kombinuje predikcie 5 zakladnych modelov s kontextovou analyzou LLM (GPT-4o-mini). Vstupy: 30 poslednych cien, predikcie modelov, technicke indikatory. temperature=0 pre reprodukovatelnost.",
        "train_start": "Zaciatok treningu",
        "train_end": "Koniec treningu",
        "test_start": "Zaciatok testu",
        "test_end": "Koniec testu",
        "window_num": "Okno #",
        "adf_title": "ADF test (stacionarita)",
        "adf_desc": "H0: rad ma jednotkovy koren (nestacionarny). p < 0.05 -> zamietnutie H0, rad je stacionarny.",
        "lb_title": "Ljung-Box test (autokorelácia)",
        "lb_desc": "H0: ziadna autokorelácia v rezidualoch. p < 0.05 -> zamietnutie H0, zistena autokorelácia.",
        "dm_title": "Diebold-Mariano test",
        "dm_desc": "H0: oba modely maju rovnaku strednu stratu. p < 0.05 -> jeden model je statisticky lepsi.",
        "statistic": "Testovacia statistika",
        "pvalue": "p-hodnota",
        "result": "Vysledok",
        "stationary": "Stacionarny",
        "not_stationary": "Nestacionarny",
        "no_autocorr": "Ziadna vyznamna autokorelácia",
        "autocorr_detected": "Zistena autokorelácia",
        "better_model": "Lepsi model",
        "significant": "Vyznamny (p < 0.05)",
        "not_significant": "Nevyznamny (p >= 0.05)",
        "explanation": "Vysvetlenie",
        "conclusion": "Zaver",
        "comparison_notes": "Poznamky k porovnaniu",
        "note_da_50": "Smerova presnost ~50% nie je lepsia ako nahodne hadanie",
        "note_da_55": "Smerova presnost >55% naznacuje zmysluplnu predikcnu schopnost",
        "note_rmse_best": "Nizsie RMSE = lepsie minimalizovanie velkosti chyb predikcie",
        "note_da_vs_naive": "Model by mal prekonat DA Naive, aby bol povazovany za skutocne lepsi",
        # ── New keys for full Slovak coverage ──
        "price_statistics": "Cenova statistika",
        "quantile_analysis": "Kvantilova analyza",
        "return_statistics": "Statistika vynosov",
        "monthly_performance": "Mesacna vykonnost",
        "positive_months": "Pozitivne mesiace",
        "avg_monthly_return": "Priemerny mesacny vynos",
        "best_month": "Najlepsi mesiac",
        "worst_month": "Najhorsi mesiac",
        "month_col": "Mesiac",
        "return_pct": "Vynos %",
        "direction": "Smer",
        "median_close": "Median Close",
        "first_close": "Prvy Close",
        "last_close": "Posledny Close",
        "total_change_pct": "Celkova zmena %",
        "mean_log_return": "Priemer log vynosu",
        "std_log_return": "Smerodajna odch. log vynosu",
        "daily_volatility": "Denna volatilita",
        "annualized_volatility": "Anualizovana volatilita",
        "skewness": "Sikmost",
        "kurtosis_excess": "Spicatost (excess)",
        "min_return": "Min. vynos",
        "max_return": "Max. vynos",
        "positive_days": "Pozitivne dni",
        "positive_days_pct": "Pozitivne dni %",
        "largest_up": "Najvacsi rast",
        "largest_down": "Najvacsi pokles",
        "jb_statistic": "Jarque-Bera statistika",
        "jb_pvalue": "Jarque-Bera p-hodnota",
        "distribution_label": "Rozdelenie",
        "dist_normal": "Normalne",
        "dist_non_normal": "Nenormalne (tazke chvosty)",
        "sw_statistic": "Shapiro-Wilk statistika",
        "sw_pvalue": "Shapiro-Wilk p-hodnota",
        "volatility_analysis": "Analyza volatility",
        "daily_vol_sigma": "Denna volatilita (sigma)",
        "ann_vol_sigma": "Anualizovana volatilita (sigma * sqrt(252))",
        "intraday_range_avg": "Priemerny intraday rozsah",
        "vol_regime": "Rezim volatility",
        "vol_low": "Nizka volatilita",
        "vol_normal": "Normalna volatilita",
        "vol_high": "Vysoka volatilita",
        "var_title": "Hodnota v riziku (VaR)",
        "confidence": "Hladina spolahlivosti",
        "var_daily": "VaR (denny)",
        "var_annualized": "VaR (anualizovany)",
        "expected_shortfall": "Ocakavany deficit (ES)",
        "drawdown_analysis": "Analyza drawdownu",
        "max_drawdown": "Maximalny drawdown",
        "max_dd_duration": "Max. trvanie drawdownu (barov)",
        "current_drawdown": "Aktualny drawdown",
        "avg_drawdown": "Priemerny drawdown",
        "rolling_vol_title": "Rolujuca 20-dnova anualizovana volatilita (%)",
        "models_used": "Modely pouzite v analyze",
        "model_type": "Typ modelu",
        "complexity_col": "Zlozitost",
        "metrics_explained": "Vysvetlenie metrik",
        "comparison_vs_naive": "Podrobne porovnanie modelov vs Naive",
        "rmse_vs_naive": "RMSE vs Naive",
        "mae_vs_naive": "MAE vs Naive",
        "da_vs_naive": "DA vs Naive",
        "verdict": "Verdikt",
        "outperforms": "Prekonava",
        "partial_result": "Ciastocne",
        "underperforms": "Nedosahuje",
        "per_model_stats": "Statistiky po modeloch",
        "error_dist_per_model": "Rozdelenie chyb podla modelov",
        "mean_error": "Priemerna chyba",
        "std_error": "Smerodajna odch. chyby",
        "min_error": "Min. chyba",
        "max_error": "Max. chyba",
        "median_error": "Median chyby",
        "error_skewness": "Sikmost chyb",
        "error_kurtosis": "Spicatost chyb",
        "mae_percentiles": "Percentily MAE podla modelov",
        "consistency_score": "Skore konzistencie modelu",
        "consistency_desc": "Nizsi koeficient variacie (CV) znamena konzistentnejsie predikcie napriec oknami.",
        "cons_excellent": "Vynikajuca",
        "cons_good": "Dobra",
        "cons_moderate": "Priemerna",
        "cons_poor": "Slaba",
        "significance_summary": "Suhrn statistickej vyznamnosti",
        "key_findings": "Klucove zistenia",
        "final_ranking": "Finalny ranking modelov",
        "methodology_title": "Metodologia",
        "beats_naive_col": "Prekonava Naive?",
        "baseline_label": "zaklad",
        "yes_label": "Ano",
        "no_label": "Nie",
        "best_label": "NAJLEPSI",
        "strong_label": "Silny",
        "moderate_label": "Priemerny",
        "weak_label": "Slaby",
        # ── Target / prediction clarity ──
        "target_info_title": "Predikcny ciel",
        "target_close": "Predikcia uzatvaracej ceny (Close) v jednotkach meny",
        "target_log_return": "Predikcia logaritmickeho vynosu (log return)",
        "target_reconstr": "Rekonstrukcia ceny: close(t+1) = close(t) * exp(r(t+1))",
        "error_on_close": "Chyba (MAE, RMSE) je merana v jednotkach ceny (napr. USD)",
        "error_on_return": "Chyba je merana v log-vynosoch",
        "error_definition": "Definicia chyby",
        # ── Data source audit ──
        "data_source_title": "Zdroj dat",
        "data_provider": "Poskytovatel",
        "data_api": "API / kniznica",
        "data_timezone": "Casove pasmo",
        "data_quality_label": "Kontrola kvality",
        "data_duplicates": "Duplicitne zaznamy",
        "data_missing": "Chybajuce obchodne dni",
        "data_cleaning": "Cistenie dat",
        "data_cleaning_desc": "Automaticka detekcia duplikatov, neplatnych barov (OHLC=0/NaN), overenie monotonnosti casovych pectiatok. Vsetky casy su v UTC.",
        # ── DA formula ──
        "da_formula_title": "Vypocet smerovej presnosti (DA)",
        "da_formula": "DA = (1/n) * sum( 1[sign(r_hat_t) = sign(r_t)] )",
        "da_baseline_text": "Nahodna uroven: ~50% (ekvivalent hadzania mincou)",
        "da_interp": "DA > 55% naznacuje zmysluplnu predikcnu schopnost; DA < 50% je horsie ako nahoda.",
        # ── Model parameters ──
        "model_params": "Parametre modelu",
        "features_used": "Pouzite premenne (features)",
        "train_val_split": "Rozdelenie train/validacia",
        "determinism": "Deterministickost",
        "ai_composition": "Zlozenie AI ansamblu",
        "ai_how": "Sposob kombinacie predikcii",
        "ai_comp_detail": "Predikcie modelov Naive, MovingAverage, ARIMA, Ridge, RandomForest + technicke indikatory (RSI, MACD, Momentum, Bollinger) + poslednych 30 cien.",
        "ai_how_detail": "LLM (GPT-4o-mini) analyzuje kontext a vracia vazenu predikciu. Fallback: jednoduchy priemer ak API nie je dostupne. temperature=0.",
        # ── Selection rule ──
        "selection_rule_title": "Pravidlo vyberu najlepsieho modelu",
        "selection_rule_desc": "Z modelov, ktorych DA >= DA(Naive), sa vyberie ten s najnizsim RMSE. Ak ziadny nespĺna tuto podmienku, vyberie sa model s najvyssou DA.",
        # ── Conclusion structure ──
        "best_by_rmse": "Najlepsi model podla RMSE",
        "best_by_mae": "Najlepsi model podla MAE",
        "best_by_da": "Najlepsi model podla DA",
        "stat_tests_conclusion": "Zavery statistickych testov",
        "limitations_title": "Obmedzenia analyzy",
        "improvements_title": "Navrhy na zlepsenie",
        # ── Top errors ──
        "top_errors_title": "Najvacsie chyby (podla MAE v oknach)",
        "error_date": "Datum",
        "error_value": "Chyba",
        # ── LB specifics ──
        "lb_residuals": "Rezidua modelu",
        "lb_lag": "Pocet lagov",
        "lb_h0_text": "H0: ziadna autokorelácia rezidualov do lagu k. p < 0.05 -> zamietnutie H0.",
        # ── Rolling performance ──
        "rolling_perf": "Rolujuca vykonnost",
        "window_label": "Okno",
        "notes_col": "Poznamky",
        # ── Model types & complexity ──
        "type_baseline": "Zaklad (baseline)",
        "type_statistical": "Statisticky",
        "type_timeseries": "Casovy rad",
        "type_ml": "Strojove ucenie",
        "type_ai_ensemble": "AI ansambel",
        "complexity_low": "Nizka",
        "complexity_medium": "Stredna",
        "complexity_high": "Vysoka",
        "complexity_very_high": "Velmi vysoka",
        # ── Model parameters detail ──
        "params_none": "Ziadne parametre",
        "features_naive": "Ziadne - pouziva len poslednu cenu",
        "deterministic_yes": "Ano (deterministicky)",
        "params_ma_window": "okno k = 20",
        "features_ma": "Poslednych 20 uzatvaracich cien",
        "params_arima": "ARIMA(p,d,q) - automaticky vyber cez AIC (pmdarima), typicky (1,1,1)",
        "features_arima": "Rad uzatvaracich cien (univariatny model)",
        "deterministic_arima": "Ano (fixny seed, rovnake data = rovnaky vysledok)",
        "train_val_desc": "Casove rozdelenie: poslednych 20% dat = validacia",
        "deterministic_ridge": "Ano (deterministicky, bez nahodnosti)",
        "deterministic_rf": "Ano (random_state=42)",
        "deterministic_ai": "Ciastocne (temperature=0, ale LLM moze varirovat)",
        "formula_label": "Vzorec",
        "baseline_da_label": "Zaklad",
        "interpretation_label": "Interpretacia",
        # ── Backtest notes ──
        "vs_naive_da": "vs Naive DA",
        "rmse_rank_col": "RMSE poradie",
        "note_lowest_rmse": "Najnizsie RMSE",
        "note_highest_da": "Najvyssia DA",
        "note_below_random": "Pod nahodnou urovnou",
        "note_beats_naive": "Prekonava Naive v oboch metrikach",
        "chart_mae_rmse": "Porovnanie MAE a RMSE",
        "chart_da_pct": "Smerova presnost (%)",
        # ── Statistical significance ──
        "lb_residuals_of": "Rezidua modelu: {}",
        "lb_lag_default": "10 (predvoleny)",
        "sig_adf_ok": "Stacionarny - log vynosy su vhodne na modelovanie.",
        "sig_adf_fail": "Nestacionarny - pozor, rad moze mat trend.",
        "sig_lb_ok": "Ziadna autokorelácia - rezidua sa spravaju ako biely sum.",
        "sig_lb_fail": "Zistena autokorelácia - model moze byt nedostatocny.",
        "sig_dm_ok": "Vyznamny - {} je statisticky lepsi.",
        "sig_dm_fail": "Nevyznamny - ziadny statisticky rozdiel medzi modelmi.",
        # ── Conclusion findings ──
        "finding_best_model": "Najlepsi celkovy model: {} (RMSE: {}, DA: {}%)",
        "finding_model_count": "Pocet testovanych modelov: {}",
        "finding_data_points": "Analyzovanych datovych bodov: {}",
        "finding_period": "Obdobie: {} — {}",
        "finding_naive_da": "DA zakladu (Naive): {}%",
        "finding_ann_vol": "Anualizovana volatilita: {}%",
        "finding_max_dd": "Maximalny drawdown: {}%",
        "finding_total_change": "Celkova zmena ceny: {}%",
        "finding_winners": "Modely prekonavajuce Naive v oboch metrikach: {}",
        "finding_no_winners": "Ziadny model neprekonava Naive sucasne v RMSE aj DA.",
        "finding_da_below_50": "Pozor: DA najlepsieho modelu ({}%) je pod 50% - horsie ako nahoda.",
        "finding_da_near_50": "DA najlepsieho modelu ({}%) je blizko nahodnej urovne - obmedzena predikcna sila.",
        # ── Conclusion: stat tests ──
        "conc_adf_stationary": "stacionarne",
        "conc_adf_nonstationary": "nestacionarne",
        "conc_adf_ok_comment": "Vhodne na modelovanie.",
        "conc_adf_fail_comment": "Pozor: moze byt potrebne dalsie diferencovanie.",
        "conc_lb_ok": "ziadna vyznamna autokorelácia v rezidualoch",
        "conc_lb_fail": "zistena autokorelácia - model moze byt nedostatocny",
        "conc_dm_ok": "vyznamny rozdiel, lepsi model: {}",
        "conc_dm_fail": "ziadny statisticky vyznamny rozdiel medzi modelmi",
        # ── Methodology ──
        "method_walkforward": "Walk-forward krizova validacia s rolujucimi oknami",
        "method_windows": "Trenovacie okno: {} barov, Testovacie okno: {} barov",
        "method_metrics": "Metriky: MAE, RMSE, Smerova presnost (DA)",
        "method_tests": "Statisticke testy: ADF (stacionarita), Ljung-Box (autokorelácia), Diebold-Mariano (porovnanie modelov)",
        "method_datasource": "Zdroj dat: Yahoo Finance (yfinance)",
        "method_generated": "Report vygenerovany: {}",
        # ── Limitations ──
        "lim_timeframe": "Analyzovany iba {} casovy ramec pre {}",
        "lim_interval": "Obmedzeny casovy interval: {} az {} ({} pozorovani)",
        "lim_no_exogenous": "Bez exogennych premennych (urokove sadzby, makroekonomicke ukazovatele, sentiment)",
        "lim_yahoo": "Yahoo Finance nie je primarny zdroj dat (agreguje z roznych poskytovatelov)",
        "lim_ai_api": "AI ansambel zavisi od externeho API (moze varirovat)",
        # ── Improvements ──
        "imp_more_pairs": "Rozsirit na viac menovych parov (GBP/USD, USD/JPY) a dlhsie obdobia",
        "imp_more_features": "Pridat dalsie features: volatility clustering (GARCH), sentiment z financnych sprav",
        "imp_deep_models": "Implementovat hlboke modely (LSTM, GRU, Transformer) pre porovnanie",
        "imp_validation": "Robustnejsia validacia: expanding window, viacnasobny walk-forward",
        "imp_exogenous": "Pridat exogenne premenne (urokove sadzby, CPI, PMI)",
        "imp_oos_verify": "Overit konzistenciu vysledkov na out-of-sample datach z ineho zdroja (ECB, FRED)",
        # ── Indicator title ──
        "current_indicator_values": "Aktualne hodnoty indikatorov (posledny bar)",
    },
}

MODEL_DESCS_MAP = {
    "Naive": "naive_desc", "MovingAverage": "ma_desc", "ARIMA": "arima_desc",
    "Ridge": "ridge_desc", "RandomForest": "rf_desc", "AIEnsemble": "ai_desc",
}


def _format_pval(p: float) -> str:
    if p < 0.001:
        return "< 0.001"
    return f"{p:.6f}"


# ─── Technical indicator calculations ──────────────────────────────────────────

def _calc_sma(series: pd.Series, window: int) -> pd.Series:
    return series.rolling(window=window, min_periods=1).mean()


def _calc_ema(series: pd.Series, span: int) -> pd.Series:
    return series.ewm(span=span, adjust=False).mean()


def _calc_rsi(series: pd.Series, period: int = 14) -> pd.Series:
    """RSI using Wilder's smoothing (EMA with com=period-1), consistent with utils/indicators.py."""
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(com=period - 1, min_periods=period).mean()
    avg_loss = loss.ewm(com=period - 1, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi.fillna(50)


def _calc_macd(series: pd.Series) -> tuple[pd.Series, pd.Series, pd.Series]:
    ema12 = series.ewm(span=12, adjust=False).mean()
    ema26 = series.ewm(span=26, adjust=False).mean()
    macd_line = ema12 - ema26
    signal_line = macd_line.ewm(span=9, adjust=False).mean()
    histogram = macd_line - signal_line
    return macd_line, signal_line, histogram


def _calc_bollinger(series: pd.Series, window: int = 20, num_std: float = 2.0):
    sma = series.rolling(window=window, min_periods=1).mean()
    std = series.rolling(window=window, min_periods=1).std().fillna(0)
    upper = sma + num_std * std
    lower = sma - num_std * std
    return sma, upper, lower


def _calc_atr(high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14) -> pd.Series:
    tr1 = high - low
    tr2 = (high - close.shift(1)).abs()
    tr3 = (low - close.shift(1)).abs()
    tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    return tr.rolling(window=period, min_periods=1).mean()


def _max_drawdown(close: pd.Series) -> tuple[float, int]:
    """Returns (max drawdown as fraction, duration in bars)."""
    cummax = close.cummax()
    safe_cummax = cummax.replace(0, np.nan)
    drawdown = (close - cummax) / safe_cummax
    drawdown = drawdown.fillna(0.0)
    max_dd = float(drawdown.min()) if len(drawdown.dropna()) > 0 else 0.0
    # duration: longest consecutive drawdown
    in_dd = drawdown < 0
    groups = (~in_dd).cumsum()
    if in_dd.any():
        dd_lens = in_dd.groupby(groups).sum()
        max_dur = int(dd_lens.max()) if len(dd_lens) > 0 else 0
    else:
        max_dur = 0
    return max_dd, max_dur


# ─── Excel generation ──────────────────────────────────────────────────────────

async def generate_excel_report(
    db: AsyncSession,
    req: GenerateReportRequest,
    precomputed_bt_resp=None,
) -> tuple[str, Path]:
    """Generate a detailed Excel report. Returns (report_id, file_path)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.chart import BarChart, LineChart, Reference, AreaChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter

    report_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    symbol = req.pair
    lang = req.language or "en"
    t = TRANSLATIONS.get(lang, TRANSLATIONS["en"])

    close = await get_close_series(db, symbol, req.timeframe.value, req.start, req.end)
    ohlc_df = await get_ohlc_dataframe(db, symbol, req.timeframe.value, req.start, req.end)

    data_start = close.index[0].strftime("%Y-%m-%d") if len(close) > 0 else "N/A"
    data_end = close.index[-1].strftime("%Y-%m-%d") if len(close) > 0 else "N/A"

    # ── Precompute common data ──
    log_returns = pd.Series(dtype=float)
    daily_vol = 0.0
    ann_vol = 0.0
    total_change_pct = 0.0
    max_dd = 0.0
    max_dd_dur = 0

    if len(close) > 1:
        log_returns = np.log(close / close.shift(1)).dropna()
        daily_vol = float(log_returns.std()) if len(log_returns) > 1 else 0
        ann_vol = daily_vol * np.sqrt(252) if daily_vol > 0 else 0
        total_change_pct = float((close.iloc[-1] / close.iloc[0] - 1) * 100)
        max_dd, max_dd_dur = _max_drawdown(close)

    # ── Styles ──
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    brand_font = Font(name="Calibri", bold=True, color=BRAND_BLUE, size=16)
    title_font = Font(name="Calibri", bold=True, size=13, color=BRAND_DARK)
    subtitle_font = Font(name="Calibri", bold=True, size=11, color=BRAND_DARK)
    section_font = Font(name="Calibri", bold=True, size=12, color=BRAND_BLUE)
    normal_font = Font(name="Calibri", size=11)
    mono_font = Font(name="Consolas", size=11)
    small_mono = Font(name="Consolas", size=10)
    desc_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    kpi_value_font = Font(name="Calibri", bold=True, size=16, color=BRAND_DARK)
    kpi_label_font = Font(name="Calibri", size=9, color="64748B")
    good_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    warn_fill = PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid")
    bad_fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
    light_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def write_kpi(ws, row, col, label, value, bg_color, num_format=None):
        """Write a KPI box (2 rows: value, label) with background."""
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        c_val = ws.cell(row=row, column=col, value=value)
        c_val.font = kpi_value_font
        c_val.fill = fill
        c_val.alignment = center_align
        c_val.border = thin_border
        if num_format:
            c_val.number_format = num_format
        c_lbl = ws.cell(row=row + 1, column=col, value=label)
        c_lbl.font = kpi_label_font
        c_lbl.fill = fill
        c_lbl.alignment = center_align
        c_lbl.border = thin_border

    def write_section_title(ws, row, col, title, end_col=6):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=col, value=title)
        c.font = section_font
        c.border = Border(bottom=Side(style="medium", color=BRAND_BLUE))
        return row + 1

    def data_cell(ws, r, c, val, is_num=False, fmt=None, font=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font or (mono_font if is_num else normal_font)
        cell.border = thin_border
        if is_num:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            cell.number_format = fmt
        return cell

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1: DASHBOARD
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = t["dashboard_sheet"]
    ws.sheet_properties.tabColor = BRAND_BLUE

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{t['title']} — {symbol}"
    ws["A1"].font = brand_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{req.timeframe.value} | {data_start} — {data_end} | Generated: {now.strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="64748B")
    ws["A2"].alignment = center_align

    # ── KPI Row ──
    kpi_row = 4
    ws.row_dimensions[kpi_row].height = 32
    ws.row_dimensions[kpi_row + 1].height = 20

    write_kpi(ws, kpi_row, 1, "DATA POINTS", len(close), KPI_BG_1)
    write_kpi(ws, kpi_row, 2, "FIRST CLOSE",
              round(float(close.iloc[0]), 5) if len(close) > 0 else "N/A", KPI_BG_2, "0.00000")
    write_kpi(ws, kpi_row, 3, "LAST CLOSE",
              round(float(close.iloc[-1]), 5) if len(close) > 0 else "N/A", KPI_BG_2, "0.00000")

    chg_color = KPI_BG_2 if total_change_pct >= 0 else KPI_BG_3
    write_kpi(ws, kpi_row, 4, "TOTAL CHANGE %",
              f"{'▲' if total_change_pct >= 0 else '▼'} {abs(total_change_pct):.3f}%", chg_color)

    write_kpi(ws, kpi_row, 5, "ANNUALIZED VOL",
              f"{ann_vol * 100:.2f}%" if ann_vol > 0 else "N/A", KPI_BG_4)

    write_kpi(ws, kpi_row, 6, "MAX DRAWDOWN",
              f"{max_dd * 100:.2f}%" if max_dd < 0 else "0.00%", KPI_BG_3 if max_dd < -0.05 else KPI_BG_5)

    write_kpi(ws, kpi_row, 7, "POSITIVE DAYS %",
              f"{float((log_returns > 0).sum() / len(log_returns) * 100):.1f}%" if len(log_returns) > 0 else "N/A",
              KPI_BG_2)

    write_kpi(ws, kpi_row, 8, "PRICE RANGE",
              round(float(close.max() - close.min()), 5) if len(close) > 0 else "N/A",
              KPI_BG_1, "0.00000")

    # ── Price mini-chart data (hidden data area for chart) ──
    chart_data_row = 8
    ws.cell(row=chart_data_row, column=1, value="Date").font = desc_font
    ws.cell(row=chart_data_row, column=2, value="Close").font = desc_font
    ws.cell(row=chart_data_row, column=3, value="SMA(20)").font = desc_font

    if len(close) > 0:
        sma20 = _calc_sma(close, 20)
        # Sample data for dashboard chart (max 200 points for readability)
        step = max(1, len(close) // 200)
        sampled = close.iloc[::step]
        sampled_sma = sma20.iloc[::step]
        for i, (idx, val) in enumerate(sampled.items()):
            ws.cell(row=chart_data_row + 1 + i, column=1, value=str(idx)[:10]).font = Font(size=8, color="94A3B8")
            ws.cell(row=chart_data_row + 1 + i, column=2, value=round(float(val), 6)).font = Font(size=8, color="94A3B8")
            sma_val = sampled_sma.iloc[i] if i < len(sampled_sma) else None
            if sma_val is not None and pd.notna(sma_val):
                ws.cell(row=chart_data_row + 1 + i, column=3, value=round(float(sma_val), 6)).font = Font(size=8, color="94A3B8")

        n_pts = len(sampled)
        if n_pts > 1:
            lc = LineChart()
            lc.title = f"{symbol} — Close Price Overview"
            lc.style = 10
            lc.width = 38
            lc.height = 16
            lc.y_axis.title = "Price"
            lc.y_axis.numFmt = "0.0000"
            d1 = Reference(ws, min_col=2, max_col=3, min_row=chart_data_row, max_row=chart_data_row + n_pts)
            cats = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + n_pts)
            lc.add_data(d1, titles_from_data=True)
            lc.set_categories(cats)
            lc.series[0].graphicalProperties.line.solidFill = "2563EB"
            lc.series[0].graphicalProperties.line.width = 18000
            if len(lc.series) > 1:
                lc.series[1].graphicalProperties.line.solidFill = "F59E0B"
                lc.series[1].graphicalProperties.line.width = 12000
                lc.series[1].graphicalProperties.line.dashStyle = "dash"
            ws.add_chart(lc, "A" + str(chart_data_row + n_pts + 2))

    # Column widths
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2: SUMMARY (enhanced)
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(t["summary_sheet"])
    ws2.sheet_properties.tabColor = "3B82F6"

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"{t['title']} — {t['summary_sheet']}"
    ws2["A1"].font = brand_font
    ws2["A1"].alignment = center_align
    ws2.row_dimensions[1].height = 40

    # ── Overview section ──
    row = 3
    row = write_section_title(ws2, row, 1, t["overview"])

    overview_data = [
        (t["pair"], symbol),
        (t["timeframe"], req.timeframe.value),
        (t["period"], f"{data_start} — {data_end}"),
        (t["data_points"], len(close)),
        (t["generated"], now.strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for label, val in overview_data:
        ws2.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
        ws2.cell(row=row, column=1).border = thin_border
        ws2.cell(row=row, column=1).fill = light_fill
        ws2.cell(row=row, column=2, value=val).font = normal_font
        ws2.cell(row=row, column=2).border = thin_border
        row += 1

    # ── Target / Prediction clarity ──
    row += 1
    row = write_section_title(ws2, row, 1, t["target_info_title"])

    target_info = [
        (t["target_info_title"], t["target_close"]),
        (t["error_definition"], t["error_on_close"]),
        (t["target_info_title"], t["target_reconstr"]),
    ]
    for label, val in target_info:
        ws2.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
        ws2.cell(row=row, column=1).border = thin_border
        ws2.cell(row=row, column=1).fill = light_fill
        ws2.cell(row=row, column=2, value=val).font = normal_font
        ws2.cell(row=row, column=2).border = thin_border
        row += 1

    # ── Data source audit ──
    row += 1
    row = write_section_title(ws2, row, 1, t["data_source_title"])

    data_src_info = [
        (t["data_provider"], "Yahoo Finance"),
        (t["data_api"], "yfinance (Python)"),
        (t["data_timezone"], "UTC"),
        (t["data_cleaning"], t["data_cleaning_desc"]),
    ]
    for label, val in data_src_info:
        ws2.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
        ws2.cell(row=row, column=1).border = thin_border
        ws2.cell(row=row, column=1).fill = light_fill
        c2 = ws2.cell(row=row, column=2, value=val)
        c2.font = normal_font
        c2.border = thin_border
        c2.alignment = wrap_align
        row += 1

    if len(close) > 0:
        # ── Price Statistics ──
        row += 1
        row = write_section_title(ws2, row, 1, t["price_statistics"])

        price_stats = [
            (t["mean"], round(float(close.mean()), 6), "0.000000"),
            (t["std"], round(float(close.std()), 6), "0.000000"),
            (t["min"], round(float(close.min()), 6), "0.000000"),
            (t["max"], round(float(close.max()), 6), "0.000000"),
            (t["range"], round(float(close.max() - close.min()), 6), "0.000000"),
            (t["median_close"], round(float(close.median()), 6), "0.000000"),
            (t["first_close"], round(float(close.iloc[0]), 6), "0.000000"),
            (t["last_close"], round(float(close.iloc[-1]), 6), "0.000000"),
            (t["total_change_pct"], round(total_change_pct, 4), "0.0000"),
        ]
        for label, val, fmt in price_stats:
            ws2.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
            ws2.cell(row=row, column=1).border = thin_border
            ws2.cell(row=row, column=1).fill = light_fill
            c = ws2.cell(row=row, column=2, value=val)
            c.font = mono_font
            c.number_format = fmt
            c.border = thin_border
            row += 1

        # ── Quantile Analysis ──
        row += 1
        row = write_section_title(ws2, row, 1, t["quantile_analysis"])

        quantiles = [0.01, 0.05, 0.10, 0.25, 0.50, 0.75, 0.90, 0.95, 0.99]
        q_header = ["Percentile"]
        q_values = ["Close Price"]
        for q in quantiles:
            q_header.append(f"P{int(q * 100)}")
            q_values.append(round(float(close.quantile(q)), 6))

        for ci, h in enumerate(q_header, 1):
            ws2.cell(row=row, column=ci, value=h)
        style_header_row(ws2, row, len(q_header))
        row += 1
        for ci, v in enumerate(q_values, 1):
            c = ws2.cell(row=row, column=ci, value=v)
            c.font = mono_font if ci > 1 else Font(name="Calibri", bold=True, size=10)
            c.border = thin_border
            if ci > 1:
                c.number_format = "0.000000"
        row += 1

        # Same for log returns
        if len(log_returns) > 0:
            q_values_ret = ["Log Return"]
            for q in quantiles:
                q_values_ret.append(round(float(log_returns.quantile(q)), 8))
            for ci, v in enumerate(q_values_ret, 1):
                c = ws2.cell(row=row, column=ci, value=v)
                c.font = mono_font if ci > 1 else Font(name="Calibri", bold=True, size=10)
                c.border = thin_border
                if ci > 1:
                    c.number_format = "0.00000000"
            row += 1

        # ── Return Statistics ──
        row += 1
        row = write_section_title(ws2, row, 1, t["return_statistics"])

        if len(log_returns) > 0:
            from scipy import stats as sp_stats

            ret_stats = [
                (t["mean_log_return"], round(float(log_returns.mean()), 8), "0.00000000"),
                (t["std_log_return"], round(float(log_returns.std()), 8), "0.00000000"),
                (t["daily_volatility"], round(daily_vol, 8), "0.00000000"),
                (t["annualized_volatility"], f"{ann_vol * 100:.4f}%", None),
                (t["skewness"], round(float(log_returns.skew()), 4), "0.0000"),
                (t["kurtosis_excess"], round(float(log_returns.kurtosis()), 4), "0.0000"),
                (t["min_return"], round(float(log_returns.min()), 8), "0.00000000"),
                (t["max_return"], round(float(log_returns.max()), 8), "0.00000000"),
                (t["positive_days"], f"{int((log_returns > 0).sum())} / {len(log_returns)}", None),
                (t["positive_days_pct"], f"{float((log_returns > 0).sum() / len(log_returns) * 100):.1f}%", None),
                (t["largest_up"], round(float(log_returns.max()), 8), "0.00000000"),
                (t["largest_down"], round(float(log_returns.min()), 8), "0.00000000"),
            ]

            # Jarque-Bera normality test
            if len(log_returns) >= 8:
                try:
                    jb_stat, jb_pval = sp_stats.jarque_bera(log_returns.values)
                    ret_stats.append((t["jb_statistic"], round(float(jb_stat), 4), "0.0000"))
                    ret_stats.append((t["jb_pvalue"], _format_pval(float(jb_pval)), None))
                    is_normal = jb_pval > 0.05
                    ret_stats.append((t["distribution_label"],
                                      t["dist_normal"] if is_normal else t["dist_non_normal"], None))
                except Exception:
                    pass

            # Shapiro-Wilk test (if small sample)
            if 8 <= len(log_returns) <= 5000:
                try:
                    sw_stat, sw_pval = sp_stats.shapiro(log_returns.values[:5000])
                    ret_stats.append((t["sw_statistic"], round(float(sw_stat), 4), "0.0000"))
                    ret_stats.append((t["sw_pvalue"], _format_pval(float(sw_pval)), None))
                except Exception:
                    pass

            for label, val, fmt in ret_stats:
                ws2.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
                ws2.cell(row=row, column=1).border = thin_border
                ws2.cell(row=row, column=1).fill = light_fill
                c = ws2.cell(row=row, column=2, value=val)
                c.font = mono_font
                if fmt:
                    c.number_format = fmt
                c.border = thin_border
                row += 1

        # ── Monthly Returns Table ──
        if len(close) > 30:
            row += 1
            row = write_section_title(ws2, row, 1, t["monthly_performance"])

            monthly = close.resample("ME").last().pct_change().dropna() * 100
            m_headers = [t["month_col"], t["return_pct"], t["direction"]]
            for ci, h in enumerate(m_headers, 1):
                ws2.cell(row=row, column=ci, value=h)
            style_header_row(ws2, row, len(m_headers))
            row += 1

            for idx_m, val_m in monthly.items():
                ws2.cell(row=row, column=1, value=idx_m.strftime("%Y-%m")).font = normal_font
                ws2.cell(row=row, column=1).border = thin_border
                c = ws2.cell(row=row, column=2, value=round(float(val_m), 4))
                c.font = Font(name="Consolas", size=11,
                              color=GOOD_GREEN if val_m > 0 else BAD_RED if val_m < 0 else BRAND_DARK)
                c.number_format = '0.0000"%"'
                c.border = thin_border
                c = ws2.cell(row=row, column=3, value="▲" if val_m > 0 else "▼" if val_m < 0 else "—")
                c.font = Font(size=11, color=GOOD_GREEN if val_m > 0 else BAD_RED if val_m < 0 else "64748B")
                c.alignment = center_align
                c.border = thin_border
                row += 1

            # Monthly summary
            pos_months = int((monthly > 0).sum())
            neg_months = int((monthly < 0).sum())
            row += 1
            ws2.cell(row=row, column=1, value=t["positive_months"]).font = subtitle_font
            ws2.cell(row=row, column=2, value=f"{pos_months} / {len(monthly)}").font = mono_font
            ws2.cell(row=row, column=2).fill = good_fill
            ws2.cell(row=row, column=2).border = thin_border
            row += 1
            ws2.cell(row=row, column=1, value=t["avg_monthly_return"]).font = subtitle_font
            ws2.cell(row=row, column=2, value=f"{float(monthly.mean()):.4f}%").font = mono_font
            ws2.cell(row=row, column=2).border = thin_border
            row += 1
            ws2.cell(row=row, column=1, value=t["best_month"]).font = subtitle_font
            ws2.cell(row=row, column=2, value=f"{float(monthly.max()):.4f}%").font = Font(name="Consolas", size=11, color=GOOD_GREEN)
            ws2.cell(row=row, column=2).border = thin_border
            row += 1
            ws2.cell(row=row, column=1, value=t["worst_month"]).font = subtitle_font
            ws2.cell(row=row, column=2, value=f"{float(monthly.min()):.4f}%").font = Font(name="Consolas", size=11, color=BAD_RED)
            ws2.cell(row=row, column=2).border = thin_border

    # ── Cumulative Returns Chart ──
    if len(log_returns) > 5:
        cum_ret = log_returns.cumsum().apply(np.exp) - 1
        cum_chart_start = row + 3
        ws2.cell(row=cum_chart_start, column=1, value="Date").font = desc_font
        ws2.cell(row=cum_chart_start, column=2, value="Cumulative Return %").font = desc_font
        step_cum = max(1, len(cum_ret) // 200)
        cum_sampled = cum_ret.iloc[::step_cum]
        for ci_idx, (idx, val) in enumerate(cum_sampled.items()):
            ws2.cell(row=cum_chart_start + 1 + ci_idx, column=1,
                     value=str(idx)[:10]).font = Font(size=8, color="94A3B8")
            ws2.cell(row=cum_chart_start + 1 + ci_idx, column=2,
                     value=round(float(val) * 100, 4)).font = Font(size=8, color="94A3B8")

        n_cum = len(cum_sampled)
        if n_cum > 1:
            cum_line = LineChart()
            cum_line.title = f"{symbol} — Cumulative Return (%)"
            cum_line.style = 10
            cum_line.width = 30
            cum_line.height = 14
            cum_d = Reference(ws2, min_col=2, min_row=cum_chart_start, max_row=cum_chart_start + n_cum)
            cum_c = Reference(ws2, min_col=1, min_row=cum_chart_start + 1, max_row=cum_chart_start + n_cum)
            cum_line.add_data(cum_d, titles_from_data=True)
            cum_line.set_categories(cum_c)
            cum_line.series[0].graphicalProperties.line.solidFill = "2563EB"
            cum_line.series[0].graphicalProperties.line.width = 18000
            ws2.add_chart(cum_line, f"D{cum_chart_start}")

    ws2.column_dimensions["A"].width = 30
    for c_idx in range(2, 12):
        ws2.column_dimensions[get_column_letter(c_idx)].width = 16

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 3: PRICE DATA (enhanced with indicators)
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet(t["data_sheet"])
    ws3.sheet_properties.tabColor = "3B82F6"

    data_headers = [
        "Date", "Open", "High", "Low", "Close", "Change", "Change %",
        "Log Return", "SMA(10)", "SMA(20)", "EMA(10)", "RSI(14)"
    ]
    for col, h in enumerate(data_headers, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header_row(ws3, 1, len(data_headers))

    if ohlc_df is not None and len(ohlc_df) > 0:
        df = ohlc_df.copy()
        if "close" in df.columns:
            df["change"] = df["close"].diff()
            df["change_pct"] = df["close"].pct_change() * 100
            df["log_return"] = np.log(df["close"] / df["close"].shift(1))
            df["sma10"] = _calc_sma(df["close"], 10)
            df["sma20"] = _calc_sma(df["close"], 20)
            df["ema10"] = _calc_ema(df["close"], 10)
            df["rsi14"] = _calc_rsi(df["close"], 14)

        for i, (idx, row_data) in enumerate(df.iterrows(), 2):
            ws3.cell(row=i, column=1, value=str(idx)[:19]).font = normal_font
            ws3.cell(row=i, column=1).border = thin_border
            for j, col_name in enumerate(["open", "high", "low", "close"], 2):
                c = data_cell(ws3, i, j, round(float(row_data.get(col_name, 0)), 6),
                              is_num=True, fmt="0.000000")
            if "change" in df.columns and pd.notna(row_data.get("change")):
                c = data_cell(ws3, i, 6, round(float(row_data["change"]), 6), is_num=True, fmt="0.000000")
                if row_data["change"] > 0:
                    c.font = Font(name="Consolas", size=11, color=GOOD_GREEN)
                elif row_data["change"] < 0:
                    c.font = Font(name="Consolas", size=11, color=BAD_RED)
            if "change_pct" in df.columns and pd.notna(row_data.get("change_pct")):
                data_cell(ws3, i, 7, round(float(row_data["change_pct"]), 4), is_num=True, fmt='0.0000"%"')
            if "log_return" in df.columns and pd.notna(row_data.get("log_return")):
                data_cell(ws3, i, 8, round(float(row_data["log_return"]), 8), is_num=True, fmt="0.00000000")
            for ci, col_name in enumerate(["sma10", "sma20", "ema10"], 9):
                if col_name in df.columns and pd.notna(row_data.get(col_name)):
                    data_cell(ws3, i, ci, round(float(row_data[col_name]), 6), is_num=True, fmt="0.000000")
            if "rsi14" in df.columns and pd.notna(row_data.get("rsi14")):
                c = data_cell(ws3, i, 12, round(float(row_data["rsi14"]), 2), is_num=True, fmt="0.00")
                rsi_val = float(row_data["rsi14"])
                if rsi_val > 70:
                    c.fill = bad_fill
                elif rsi_val < 30:
                    c.fill = good_fill

        # Price + SMA chart
        if len(df) > 1:
            chart = LineChart()
            chart.title = f"{symbol} Close Price with SMA"
            chart.style = 10
            chart.y_axis.title = "Price"
            chart.y_axis.numFmt = "0.0000"
            chart.width = 32
            chart.height = 16
            data_ref = Reference(ws3, min_col=5, max_col=5, min_row=1, max_row=len(df) + 1)
            sma_ref = Reference(ws3, min_col=10, max_col=10, min_row=1, max_row=len(df) + 1)
            dates_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(df) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.add_data(sma_ref, titles_from_data=True)
            chart.set_categories(dates_ref)
            chart.series[0].graphicalProperties.line.solidFill = "2563EB"
            chart.series[0].graphicalProperties.line.width = 18000
            if len(chart.series) > 1:
                chart.series[1].graphicalProperties.line.solidFill = "F59E0B"
                chart.series[1].graphicalProperties.line.width = 12000
                chart.series[1].graphicalProperties.line.dashStyle = "dash"
            ws3.add_chart(chart, f"N2")

            # Returns area chart
            ret_chart = AreaChart()
            ret_chart.title = "Log Returns"
            ret_chart.style = 10
            ret_chart.width = 32
            ret_chart.height = 12
            ret_data = Reference(ws3, min_col=8, min_row=1, max_row=len(df) + 1)
            ret_chart.add_data(ret_data, titles_from_data=True)
            ret_chart.set_categories(dates_ref)
            ret_chart.series[0].graphicalProperties.solidFill = "93C5FD"
            ret_chart.series[0].graphicalProperties.line.solidFill = "2563EB"
            ret_chart.series[0].graphicalProperties.line.width = 8000
            ws3.add_chart(ret_chart, f"N20")

    for i in range(1, 13):
        ws3.column_dimensions[get_column_letter(i)].width = 16

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 4: TECHNICAL INDICATORS (NEW)
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet(t["indicators_sheet"])
    ws4.sheet_properties.tabColor = ACCENT_TEAL

    ws4.merge_cells("A1:J1")
    ws4["A1"] = f"{t['indicators_sheet']}"
    ws4["A1"].font = brand_font
    ws4["A1"].alignment = center_align
    ws4.row_dimensions[1].height = 40

    if ohlc_df is not None and len(ohlc_df) > 0:
        df_ind = ohlc_df.copy()
        c_close = df_ind["close"] if "close" in df_ind.columns else pd.Series(dtype=float)
        c_high = df_ind["high"] if "high" in df_ind.columns else c_close
        c_low = df_ind["low"] if "low" in df_ind.columns else c_close

        if len(c_close) > 0:
            sma10 = _calc_sma(c_close, 10)
            sma20 = _calc_sma(c_close, 20)
            sma50 = _calc_sma(c_close, 50)
            ema10 = _calc_ema(c_close, 10)
            ema20 = _calc_ema(c_close, 20)
            rsi = _calc_rsi(c_close, 14)
            macd_line, signal_line, macd_hist = _calc_macd(c_close)
            bb_mid, bb_upper, bb_lower = _calc_bollinger(c_close)
            atr = _calc_atr(c_high, c_low, c_close, 14)

            ind_headers = [
                "Date", "Close", "SMA(10)", "SMA(20)", "SMA(50)",
                "EMA(10)", "EMA(20)", "RSI(14)", "MACD", "Signal",
                "MACD Hist", "BB Upper", "BB Mid", "BB Lower", "ATR(14)"
            ]
            for ci, h in enumerate(ind_headers, 1):
                ws4.cell(row=3, column=ci, value=h)
            style_header_row(ws4, 3, len(ind_headers))

            for i, idx in enumerate(df_ind.index, 4):
                pos = i - 4
                ws4.cell(row=i, column=1, value=str(idx)[:10]).font = Font(size=9)
                ws4.cell(row=i, column=1).border = thin_border
                data_cell(ws4, i, 2, round(float(c_close.iloc[pos]), 6), True, "0.000000")
                data_cell(ws4, i, 3, round(float(sma10.iloc[pos]), 6), True, "0.000000")
                data_cell(ws4, i, 4, round(float(sma20.iloc[pos]), 6), True, "0.000000")
                data_cell(ws4, i, 5, round(float(sma50.iloc[pos]), 6) if pos >= 49 else "", True, "0.000000")
                data_cell(ws4, i, 6, round(float(ema10.iloc[pos]), 6), True, "0.000000")
                data_cell(ws4, i, 7, round(float(ema20.iloc[pos]), 6), True, "0.000000")

                rsi_val = round(float(rsi.iloc[pos]), 2)
                c_r = data_cell(ws4, i, 8, rsi_val, True, "0.00")
                if rsi_val > 70:
                    c_r.fill = bad_fill
                    c_r.font = Font(name="Consolas", size=10, color=BAD_RED, bold=True)
                elif rsi_val < 30:
                    c_r.fill = good_fill
                    c_r.font = Font(name="Consolas", size=10, color=GOOD_GREEN, bold=True)

                data_cell(ws4, i, 9, round(float(macd_line.iloc[pos]), 8), True, "0.00000000")
                data_cell(ws4, i, 10, round(float(signal_line.iloc[pos]), 8), True, "0.00000000")

                hist_val = float(macd_hist.iloc[pos])
                c_h = data_cell(ws4, i, 11, round(hist_val, 8), True, "0.00000000")
                if hist_val > 0:
                    c_h.font = Font(name="Consolas", size=10, color=GOOD_GREEN)
                elif hist_val < 0:
                    c_h.font = Font(name="Consolas", size=10, color=BAD_RED)

                data_cell(ws4, i, 12, round(float(bb_upper.iloc[pos]), 6), True, "0.000000")
                data_cell(ws4, i, 13, round(float(bb_mid.iloc[pos]), 6), True, "0.000000")
                data_cell(ws4, i, 14, round(float(bb_lower.iloc[pos]), 6), True, "0.000000")
                data_cell(ws4, i, 15, round(float(atr.iloc[pos]), 6), True, "0.000000")

            n_rows = len(df_ind)
            # RSI chart
            rsi_chart = LineChart()
            rsi_chart.title = "RSI(14)"
            rsi_chart.style = 10
            rsi_chart.width = 30
            rsi_chart.height = 12
            rsi_chart.y_axis.scaling.min = 0
            rsi_chart.y_axis.scaling.max = 100
            rsi_data = Reference(ws4, min_col=8, min_row=3, max_row=3 + n_rows)
            rsi_cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + n_rows)
            rsi_chart.add_data(rsi_data, titles_from_data=True)
            rsi_chart.set_categories(rsi_cats)
            rsi_chart.series[0].graphicalProperties.line.solidFill = ACCENT_PURPLE
            rsi_chart.series[0].graphicalProperties.line.width = 15000
            ws4.add_chart(rsi_chart, f"Q3")

            # MACD chart
            macd_chart = BarChart()
            macd_chart.type = "col"
            macd_chart.title = "MACD Histogram"
            macd_chart.style = 10
            macd_chart.width = 30
            macd_chart.height = 12
            macd_data = Reference(ws4, min_col=11, min_row=3, max_row=3 + n_rows)
            macd_cats = Reference(ws4, min_col=1, min_row=4, max_row=3 + n_rows)
            macd_chart.add_data(macd_data, titles_from_data=True)
            macd_chart.set_categories(macd_cats)
            macd_chart.series[0].graphicalProperties.solidFill = BRAND_BLUE
            ws4.add_chart(macd_chart, f"Q20")

            # Bollinger Bands chart (Close + BB Upper + BB Mid + BB Lower)
            bb_chart = LineChart()
            bb_chart.title = f"{symbol} — Bollinger Bands (20, 2σ)"
            bb_chart.style = 10
            bb_chart.width = 30
            bb_chart.height = 14
            bb_chart.y_axis.title = "Price"
            bb_chart.y_axis.numFmt = "0.0000"
            bb_close_ref = Reference(ws4, min_col=2, min_row=3, max_row=3 + n_rows)
            bb_upper_ref = Reference(ws4, min_col=12, min_row=3, max_row=3 + n_rows)
            bb_mid_ref = Reference(ws4, min_col=13, min_row=3, max_row=3 + n_rows)
            bb_lower_ref = Reference(ws4, min_col=14, min_row=3, max_row=3 + n_rows)
            bb_cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=3 + n_rows)
            bb_chart.add_data(bb_close_ref, titles_from_data=True)
            bb_chart.add_data(bb_upper_ref, titles_from_data=True)
            bb_chart.add_data(bb_mid_ref, titles_from_data=True)
            bb_chart.add_data(bb_lower_ref, titles_from_data=True)
            bb_chart.set_categories(bb_cats_ref)
            bb_chart.series[0].graphicalProperties.line.solidFill = "2563EB"
            bb_chart.series[0].graphicalProperties.line.width = 18000
            if len(bb_chart.series) > 1:
                bb_chart.series[1].graphicalProperties.line.solidFill = "94A3B8"
                bb_chart.series[1].graphicalProperties.line.width = 10000
                bb_chart.series[1].graphicalProperties.line.dashStyle = "dot"
            if len(bb_chart.series) > 2:
                bb_chart.series[2].graphicalProperties.line.solidFill = "F59E0B"
                bb_chart.series[2].graphicalProperties.line.width = 12000
                bb_chart.series[2].graphicalProperties.line.dashStyle = "dash"
            if len(bb_chart.series) > 3:
                bb_chart.series[3].graphicalProperties.line.solidFill = "94A3B8"
                bb_chart.series[3].graphicalProperties.line.width = 10000
                bb_chart.series[3].graphicalProperties.line.dashStyle = "dot"
            ws4.add_chart(bb_chart, f"Q37")

            # EMA Crossover chart (Close + EMA10 + EMA20)
            ema_chart = LineChart()
            ema_chart.title = f"{symbol} — EMA Crossover (10/20)"
            ema_chart.style = 10
            ema_chart.width = 30
            ema_chart.height = 14
            ema_chart.y_axis.title = "Price"
            ema_chart.y_axis.numFmt = "0.0000"
            ema_close_ref = Reference(ws4, min_col=2, min_row=3, max_row=3 + n_rows)
            ema10_ref = Reference(ws4, min_col=6, min_row=3, max_row=3 + n_rows)
            ema20_ref = Reference(ws4, min_col=7, min_row=3, max_row=3 + n_rows)
            ema_cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=3 + n_rows)
            ema_chart.add_data(ema_close_ref, titles_from_data=True)
            ema_chart.add_data(ema10_ref, titles_from_data=True)
            ema_chart.add_data(ema20_ref, titles_from_data=True)
            ema_chart.set_categories(ema_cats_ref)
            ema_chart.series[0].graphicalProperties.line.solidFill = "2563EB"
            ema_chart.series[0].graphicalProperties.line.width = 16000
            if len(ema_chart.series) > 1:
                ema_chart.series[1].graphicalProperties.line.solidFill = GOOD_GREEN
                ema_chart.series[1].graphicalProperties.line.width = 14000
            if len(ema_chart.series) > 2:
                ema_chart.series[2].graphicalProperties.line.solidFill = BAD_RED
                ema_chart.series[2].graphicalProperties.line.width = 14000
            ws4.add_chart(ema_chart, f"Q54")

            # Indicator summary table
            sum_row = 4 + n_rows + 2
            ws4.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6)
            ws4.cell(row=sum_row, column=1, value=t["current_indicator_values"]).font = title_font
            sum_row += 1

            latest = {
                "Close": round(float(c_close.iloc[-1]), 6),
                "SMA(10)": round(float(sma10.iloc[-1]), 6),
                "SMA(20)": round(float(sma20.iloc[-1]), 6),
                "SMA(50)": round(float(sma50.iloc[-1]), 6) if len(c_close) >= 50 else "N/A",
                "EMA(10)": round(float(ema10.iloc[-1]), 6),
                "EMA(20)": round(float(ema20.iloc[-1]), 6),
                "RSI(14)": round(float(rsi.iloc[-1]), 2),
                "MACD Line": round(float(macd_line.iloc[-1]), 8),
                "Signal Line": round(float(signal_line.iloc[-1]), 8),
                "MACD Histogram": round(float(macd_hist.iloc[-1]), 8),
                "BB Upper": round(float(bb_upper.iloc[-1]), 6),
                "BB Lower": round(float(bb_lower.iloc[-1]), 6),
                "ATR(14)": round(float(atr.iloc[-1]), 6),
            }
            for lbl, val in latest.items():
                ws4.cell(row=sum_row, column=1, value=lbl).font = subtitle_font
                ws4.cell(row=sum_row, column=1).border = thin_border
                ws4.cell(row=sum_row, column=1).fill = light_fill
                c = ws4.cell(row=sum_row, column=2, value=val)
                c.font = mono_font
                c.border = thin_border
                sum_row += 1

    for ci in range(1, 16):
        ws4.column_dimensions[get_column_letter(ci)].width = 15

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 5: RISK ANALYSIS (NEW)
    # ═══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet(t["risk_sheet"])
    ws5.sheet_properties.tabColor = BAD_RED

    ws5.merge_cells("A1:F1")
    ws5["A1"] = f"{t['risk_sheet']}"
    ws5["A1"].font = brand_font
    ws5["A1"].alignment = center_align
    ws5.row_dimensions[1].height = 40

    if len(log_returns) > 10:
        row = 3
        row = write_section_title(ws5, row, 1, t["volatility_analysis"])

        vol_stats = [
            (t["daily_vol_sigma"], f"{daily_vol * 100:.6f}%"),
            (t["ann_vol_sigma"], f"{ann_vol * 100:.4f}%"),
            (t["intraday_range_avg"], f"{float((ohlc_df['high'] - ohlc_df['low']).mean()):.6f}" if ohlc_df is not None and 'high' in ohlc_df.columns else "N/A"),
        ]

        # Volatility regime
        if ann_vol < 0.05:
            regime = t["vol_low"]
            regime_color = GOOD_GREEN
        elif ann_vol < 0.15:
            regime = t["vol_normal"]
            regime_color = WARN_AMBER
        else:
            regime = t["vol_high"]
            regime_color = BAD_RED
        vol_stats.append((t["vol_regime"], regime))

        for label, val in vol_stats:
            ws5.cell(row=row, column=1, value=label).font = subtitle_font
            ws5.cell(row=row, column=1).border = thin_border
            ws5.cell(row=row, column=1).fill = light_fill
            c = ws5.cell(row=row, column=2, value=val)
            c.font = mono_font
            c.border = thin_border
            if "Regime" in label:
                c.font = Font(name="Calibri", bold=True, size=11, color=regime_color)
            row += 1

        # ── Value at Risk ──
        row += 1
        row = write_section_title(ws5, row, 1, t["var_title"])

        var_levels = [0.01, 0.025, 0.05, 0.10]
        var_headers = [t["confidence"], t["var_daily"], t["var_annualized"], t["expected_shortfall"]]
        for ci, h in enumerate(var_headers, 1):
            ws5.cell(row=row, column=ci, value=h)
        style_header_row(ws5, row, len(var_headers))
        row += 1

        for alpha in var_levels:
            confidence = f"{(1 - alpha) * 100:.1f}%"
            var_daily = float(log_returns.quantile(alpha))
            var_annual = var_daily * np.sqrt(252)
            # Expected Shortfall (CVaR)
            es = float(log_returns[log_returns <= var_daily].mean()) if (log_returns <= var_daily).any() else var_daily

            ws5.cell(row=row, column=1, value=confidence).font = subtitle_font
            ws5.cell(row=row, column=1).alignment = center_align
            ws5.cell(row=row, column=1).border = thin_border

            c = ws5.cell(row=row, column=2, value=f"{var_daily * 100:.6f}%")
            c.font = Font(name="Consolas", size=11, color=BAD_RED)
            c.border = thin_border

            c = ws5.cell(row=row, column=3, value=f"{var_annual * 100:.4f}%")
            c.font = Font(name="Consolas", size=11, color=BAD_RED)
            c.border = thin_border

            c = ws5.cell(row=row, column=4, value=f"{es * 100:.6f}%")
            c.font = Font(name="Consolas", size=11, color=BAD_RED)
            c.border = thin_border
            row += 1

        # ── Drawdown Analysis ──
        row += 1
        row = write_section_title(ws5, row, 1, t["drawdown_analysis"])

        dd_stats = [
            (t["max_drawdown"], f"{max_dd * 100:.4f}%"),
            (t["max_dd_duration"], max_dd_dur),
        ]

        cummax = close.cummax()
        dd_series = (close - cummax) / cummax.replace(0, np.nan)
        dd_series = dd_series.fillna(0.0)
        dd_stats.append((t["current_drawdown"],
                         f"{float(dd_series.iloc[-1]) * 100:.4f}%"))
        dd_stats.append((t["avg_drawdown"],
                         f"{float(dd_series[dd_series < 0].mean()) * 100:.4f}%" if (dd_series < 0).any() else "0.00%"))

        for label, val in dd_stats:
            ws5.cell(row=row, column=1, value=label).font = subtitle_font
            ws5.cell(row=row, column=1).border = thin_border
            ws5.cell(row=row, column=1).fill = light_fill
            c = ws5.cell(row=row, column=2, value=val)
            c.font = mono_font
            c.border = thin_border
            row += 1

        # Drawdown chart data
        row += 1
        dd_chart_start = row
        ws5.cell(row=row, column=1, value="Date").font = desc_font
        ws5.cell(row=row, column=2, value="Drawdown %").font = desc_font
        step_dd = max(1, len(dd_series) // 200)
        dd_sampled = dd_series.iloc[::step_dd]
        for di, (idx, val) in enumerate(dd_sampled.items()):
            ws5.cell(row=row + 1 + di, column=1, value=str(idx)[:10]).font = Font(size=8, color="94A3B8")
            ws5.cell(row=row + 1 + di, column=2, value=round(float(val) * 100, 4)).font = Font(size=8, color="94A3B8")

        n_dd = len(dd_sampled)
        if n_dd > 1:
            dd_chart = AreaChart()
            dd_chart.title = "Drawdown Over Time"
            dd_chart.style = 10
            dd_chart.width = 30
            dd_chart.height = 14
            dd_d = Reference(ws5, min_col=2, min_row=dd_chart_start, max_row=dd_chart_start + n_dd)
            dd_c = Reference(ws5, min_col=1, min_row=dd_chart_start + 1, max_row=dd_chart_start + n_dd)
            dd_chart.add_data(dd_d, titles_from_data=True)
            dd_chart.set_categories(dd_c)
            dd_chart.series[0].graphicalProperties.solidFill = "FCA5A5"
            dd_chart.series[0].graphicalProperties.line.solidFill = BAD_RED
            dd_chart.series[0].graphicalProperties.line.width = 12000
            ws5.add_chart(dd_chart, f"D{dd_chart_start}")

        # ── Rolling Volatility ──
        if len(log_returns) > 20:
            rv_start = dd_chart_start + n_dd + 18
            ws5.cell(row=rv_start, column=1, value="Date").font = desc_font
            ws5.cell(row=rv_start, column=2, value="Rolling Vol (20d)").font = desc_font
            roll_vol = log_returns.rolling(20).std() * np.sqrt(252) * 100
            rv_sampled = roll_vol.dropna().iloc[::step_dd]
            for ri, (idx, val) in enumerate(rv_sampled.items()):
                ws5.cell(row=rv_start + 1 + ri, column=1, value=str(idx)[:10]).font = Font(size=8, color="94A3B8")
                ws5.cell(row=rv_start + 1 + ri, column=2, value=round(float(val), 4)).font = Font(size=8, color="94A3B8")

            n_rv = len(rv_sampled)
            if n_rv > 1:
                rv_chart = LineChart()
                rv_chart.title = "Rolling 20-Day Annualized Volatility (%)"
                rv_chart.style = 10
                rv_chart.width = 30
                rv_chart.height = 14
                rv_d = Reference(ws5, min_col=2, min_row=rv_start, max_row=rv_start + n_rv)
                rv_c = Reference(ws5, min_col=1, min_row=rv_start + 1, max_row=rv_start + n_rv)
                rv_chart.add_data(rv_d, titles_from_data=True)
                rv_chart.set_categories(rv_c)
                rv_chart.series[0].graphicalProperties.line.solidFill = ACCENT_PURPLE
                rv_chart.series[0].graphicalProperties.line.width = 15000
                ws5.add_chart(rv_chart, f"D{rv_start}")

    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 22
    ws5.column_dimensions["C"].width = 22
    ws5.column_dimensions["D"].width = 22

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 6: MODEL DETAILS (enhanced)
    # ═══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet(t["models_sheet"])
    ws6.sheet_properties.tabColor = "8B5CF6"

    ws6.merge_cells("A1:D1")
    ws6["A1"] = f"{t['models_sheet']}"
    ws6["A1"].font = brand_font
    ws6["A1"].alignment = center_align
    ws6.row_dimensions[1].height = 40

    row = 3
    row = write_section_title(ws6, row, 1, t["models_used"], end_col=4)

    model_headers = [t["model_col"], t["explanation"], t["model_type"], t["complexity_col"]]
    for ci, h in enumerate(model_headers, 1):
        ws6.cell(row=row, column=ci, value=h)
    style_header_row(ws6, row, 4)
    row += 1

    model_types = {
        "Naive": (t["type_baseline"], t["complexity_low"]),
        "MovingAverage": (t["type_statistical"], t["complexity_low"]),
        "ARIMA": (t["type_timeseries"], t["complexity_medium"]),
        "Ridge": (t["type_ml"], t["complexity_medium"]),
        "RandomForest": (t["type_ml"], t["complexity_high"]),
        "AIEnsemble": (t["type_ai_ensemble"], t["complexity_very_high"]),
    }
    # Language-independent complexity levels for color-coding
    _COMPLEXITY_LEVELS = {
        "Naive": 1, "MovingAverage": 1, "ARIMA": 2,
        "Ridge": 2, "RandomForest": 3, "AIEnsemble": 4,
    }

    model_names = [m.value for m in req.models]
    for model_name in model_names:
        is_ai = model_name == "AIEnsemble"
        ws6.cell(row=row, column=1, value=model_name).font = Font(
            name="Calibri", bold=True, size=11,
            color=BRAND_BLUE if is_ai else BRAND_DARK)
        ws6.cell(row=row, column=1).border = thin_border
        if is_ai:
            ws6.cell(row=row, column=1).fill = blue_fill

        desc_key = MODEL_DESCS_MAP.get(model_name, "")
        ws6.cell(row=row, column=2, value=t.get(desc_key, model_name)).font = normal_font
        ws6.cell(row=row, column=2).border = thin_border
        ws6.cell(row=row, column=2).alignment = wrap_align

        m_type, m_complexity = model_types.get(model_name, ("Other", "N/A"))
        ws6.cell(row=row, column=3, value=m_type).font = normal_font
        ws6.cell(row=row, column=3).border = thin_border
        ws6.cell(row=row, column=3).alignment = center_align

        c_comp = ws6.cell(row=row, column=4, value=m_complexity)
        _clevel = _COMPLEXITY_LEVELS.get(model_name, 0)
        c_comp.font = Font(name="Calibri", bold=True, size=10,
                           color=GOOD_GREEN if _clevel <= 1 else
                           WARN_AMBER if _clevel == 2 else BAD_RED)
        c_comp.alignment = center_align
        c_comp.border = thin_border
        row += 1

    # Metric explanations
    row += 1
    row = write_section_title(ws6, row, 1, t["metrics_explained"], end_col=4)

    for label, desc in [
        (t["mae_col"], t["mae_desc"]),
        (t["rmse_col"], t["rmse_desc"]),
        (t["da_col"], t["da_desc"]),
    ]:
        ws6.cell(row=row, column=1, value=label).font = subtitle_font
        ws6.cell(row=row, column=1).border = thin_border
        ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws6.cell(row=row, column=2, value=desc).font = desc_font
        ws6.cell(row=row, column=2).alignment = wrap_align
        ws6.cell(row=row, column=2).border = thin_border
        row += 1

    # Comparison notes
    row += 1
    row = write_section_title(ws6, row, 1, t["comparison_notes"], end_col=4)
    for note_key in ["note_da_50", "note_da_55", "note_rmse_best", "note_da_vs_naive"]:
        ws6.cell(row=row, column=1, value="•").font = normal_font
        ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws6.cell(row=row, column=2, value=t[note_key]).font = desc_font
        ws6.cell(row=row, column=2).alignment = wrap_align
        row += 1

    # ── Model parameters detail ──
    row += 1
    row = write_section_title(ws6, row, 1, t["model_params"], end_col=4)

    model_params_data = {
        "Naive": [
            (t["model_params"], t["params_none"]),
            (t["features_used"], t["features_naive"]),
            (t["determinism"], t["deterministic_yes"]),
        ],
        "MovingAverage": [
            (t["model_params"], t["params_ma_window"]),
            (t["features_used"], t["features_ma"]),
            (t["determinism"], t["deterministic_yes"]),
        ],
        "ARIMA": [
            (t["model_params"], t["params_arima"]),
            (t["features_used"], t["features_arima"]),
            (t["determinism"], t["deterministic_arima"]),
        ],
        "Ridge": [
            (t["model_params"], "alpha=1.0, Pipeline: StandardScaler + Ridge"),
            (t["features_used"], "5 lag returns, SMA(20), EMA(20), RSI(14), MACD(12,26,9)"),
            (t["train_val_split"], t["train_val_desc"]),
            (t["determinism"], t["deterministic_ridge"]),
        ],
        "RandomForest": [
            (t["model_params"], "n_estimators=100, max_depth=10, random_state=42"),
            (t["features_used"], "5 lag returns, SMA(20), EMA(20), RSI(14), MACD(12,26,9)"),
            (t["train_val_split"], t["train_val_desc"]),
            (t["determinism"], t["deterministic_rf"]),
        ],
        "AIEnsemble": [
            (t["ai_composition"], t["ai_comp_detail"]),
            (t["ai_how"], t["ai_how_detail"]),
            (t["determinism"], t["deterministic_ai"]),
        ],
    }

    for model_name in [m.value for m in req.models]:
        params = model_params_data.get(model_name, [])
        if not params:
            continue
        ws6.cell(row=row, column=1, value=model_name).font = Font(name="Calibri", bold=True, size=11, color=BRAND_BLUE)
        ws6.cell(row=row, column=1).border = thin_border
        row += 1
        for label, val in params:
            ws6.cell(row=row, column=1, value=f"  {label}").font = Font(name="Calibri", size=10, color="64748B")
            ws6.cell(row=row, column=1).border = thin_border
            ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws6.cell(row=row, column=2, value=val).font = Font(name="Calibri", size=10)
            ws6.cell(row=row, column=2).alignment = wrap_align
            ws6.cell(row=row, column=2).border = thin_border
            row += 1

    # ── DA formula ──
    row += 1
    row = write_section_title(ws6, row, 1, t["da_formula_title"], end_col=4)
    for label, val in [
        (t["formula_label"], t["da_formula"]),
        (t["baseline_da_label"], t["da_baseline_text"]),
        (t["interpretation_label"], t["da_interp"]),
    ]:
        ws6.cell(row=row, column=1, value=label).font = subtitle_font
        ws6.cell(row=row, column=1).border = thin_border
        ws6.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws6.cell(row=row, column=2, value=val).font = desc_font
        ws6.cell(row=row, column=2).alignment = wrap_align
        ws6.cell(row=row, column=2).border = thin_border
        row += 1

    ws6.column_dimensions["A"].width = 28
    ws6.column_dimensions["B"].width = 55
    ws6.column_dimensions["C"].width = 18
    ws6.column_dimensions["D"].width = 16

    # ═══════════════════════════════════════════════════════════════════════════
    # BACKTEST SHEETS (7, 8, 9, 10, 11)
    # ═══════════════════════════════════════════════════════════════════════════
    backtest_data = None
    tests_data = None
    sorted_results = []
    naive_da = 0.5
    genuine_winners = []

    if req.models and len(close) >= 60:
        if precomputed_bt_resp is not None:
            bt_resp = precomputed_bt_resp
        else:
            from app.services.backtest_service import run_backtest
            from app.schemas.schemas import BacktestRequest

            bt_req = BacktestRequest(
                pair=symbol,
                timeframe=req.timeframe,
                start=req.start,
                end=req.end,
                models=req.models,
                windowTrainDays=max(30, len(close) // 3),
                windowTestDays=max(5, len(close) // 10),
                stepDays=max(5, len(close) // 10),
            )
            bt_resp = await run_backtest(db, bt_req)

        sorted_results = sorted(bt_resp.results, key=lambda r: r.metrics.rmse)
        backtest_data = sorted_results
        tests_data = bt_resp.tests

        for r in sorted_results:
            if r.model.value == "Naive":
                naive_da = r.metrics.directionalAccuracy
                break

        best_rmse_idx = 0
        best_da_idx = max(range(len(sorted_results)),
                          key=lambda i: sorted_results[i].metrics.directionalAccuracy)

        best_rmse_model = sorted_results[best_rmse_idx]
        if best_rmse_model.metrics.directionalAccuracy >= naive_da:
            overall_best_idx = best_rmse_idx
        else:
            overall_best_idx = best_da_idx

        da_rmse_conflict = (best_rmse_idx != best_da_idx)

        naive_result = None
        for r in sorted_results:
            if r.model.value == "Naive":
                naive_result = r
                break

        if naive_result:
            for r in sorted_results:
                if r.model.value == "Naive":
                    continue
                if (r.metrics.rmse < naive_result.metrics.rmse and
                    r.metrics.directionalAccuracy > naive_result.metrics.directionalAccuracy):
                    genuine_winners.append(r.model.value)

        # ── Sheet 7: BACKTEST RESULTS (enhanced) ──
        ws7 = wb.create_sheet(t["backtest_sheet"])
        ws7.sheet_properties.tabColor = "F59E0B"

        ws7.merge_cells("A1:H1")
        ws7["A1"] = f"{t['backtest_sheet']}"
        ws7["A1"].font = brand_font
        ws7["A1"].alignment = center_align
        ws7.row_dimensions[1].height = 40

        # ── Selection rule explanation ──
        row = 3
        row = write_section_title(ws7, row, 1, t["selection_rule_title"], end_col=8)
        ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws7.cell(row=row, column=1, value=t["selection_rule_desc"]).font = desc_font
        ws7.cell(row=row, column=1).alignment = wrap_align
        row += 2

        # Main results table
        bt_headers = [t["model_col"], t["mae_col"], t["rmse_col"], t["da_col"],
                       t["rank_col"], t["vs_naive_da"], t["rmse_rank_col"], t["notes_col"]]
        for ci, h in enumerate(bt_headers, 1):
            ws7.cell(row=row, column=ci, value=h)
        style_header_row(ws7, row, len(bt_headers))
        data_start_row = row + 1

        for i, r in enumerate(sorted_results):
            row_num = data_start_row + i
            da = r.metrics.directionalAccuracy
            da_pct = da * 100
            is_overall = (i == overall_best_idx)

            model_label = r.model.value
            if is_overall:
                model_label = f"[*] {model_label}"

            c = ws7.cell(row=row_num, column=1, value=model_label)
            c.font = Font(name="Calibri", bold=True, size=11,
                          color=BRAND_BLUE if is_overall else BRAND_DARK)
            c.border = thin_border
            if is_overall:
                c.fill = blue_fill

            c = data_cell(ws7, row_num, 2, round(r.metrics.mae, 6), True, "0.000000")
            c = data_cell(ws7, row_num, 3, round(r.metrics.rmse, 6), True, "0.000000")
            if i == best_rmse_idx:
                c.font = Font(name="Consolas", size=11, bold=True, color=GOOD_GREEN)

            c = data_cell(ws7, row_num, 4, round(da_pct, 1), True, '0.0"%"')
            if da_pct >= 55:
                c.fill = good_fill
                c.font = Font(name="Consolas", size=11, color=GOOD_GREEN, bold=True)
            elif da_pct < 50:
                c.fill = bad_fill
                c.font = Font(name="Consolas", size=11, color=BAD_RED)
            else:
                c.fill = warn_fill

            c = data_cell(ws7, row_num, 5, i + 1, True)
            c.alignment = center_align

            if r.model.value != "Naive":
                diff = da - naive_da
                diff_pct = diff * 100
                vs_text = f"{'+' if diff_pct >= 0 else ''}{diff_pct:.1f}%"
                c = ws7.cell(row=row_num, column=6, value=vs_text)
                c.font = Font(name="Consolas", size=11,
                              color=GOOD_GREEN if diff > 0 else BAD_RED if diff < 0 else "64748B")
                c.alignment = center_align
                c.border = thin_border
            else:
                c = ws7.cell(row=row_num, column=6, value=t["baseline_label"])
                c.font = desc_font
                c.alignment = center_align
                c.border = thin_border

            # RMSE rank
            rmse_ranked = sorted(range(len(sorted_results)), key=lambda x: sorted_results[x].metrics.rmse)
            rmse_rank = rmse_ranked.index(i) + 1
            data_cell(ws7, row_num, 7, rmse_rank, True)
            ws7.cell(row=row_num, column=7).alignment = center_align

            # Notes
            notes = []
            if is_overall:
                notes.append(t["best_label"])
            if i == best_rmse_idx and i != overall_best_idx:
                notes.append(t["note_lowest_rmse"])
            if i == best_da_idx and i != overall_best_idx:
                notes.append(t["note_highest_da"])
            if da_pct < 50:
                notes.append(t["note_below_random"])
            if r.model.value in genuine_winners:
                notes.append(t["note_beats_naive"])
            c = ws7.cell(row=row_num, column=8, value="; ".join(notes) if notes else "")
            c.font = desc_font
            c.border = thin_border
            c.alignment = wrap_align

        # ── Detailed Comparison Table ──
        comp_row = data_start_row + len(sorted_results) + 2
        ws7.merge_cells(start_row=comp_row, start_column=1, end_row=comp_row, end_column=8)
        ws7.cell(row=comp_row, column=1, value=t["comparison_vs_naive"]).font = section_font
        comp_row += 1

        comp_headers = [t["model_col"], "MAE", "RMSE", "DA %",
                        t["rmse_vs_naive"], t["mae_vs_naive"], t["da_vs_naive"], t["verdict"]]
        for ci, h in enumerate(comp_headers, 1):
            ws7.cell(row=comp_row, column=ci, value=h)
        style_header_row(ws7, comp_row, len(comp_headers))
        comp_row += 1

        for r in sorted_results:
            da_pct = r.metrics.directionalAccuracy * 100
            ws7.cell(row=comp_row, column=1, value=r.model.value).font = Font(name="Calibri", bold=True, size=10)
            ws7.cell(row=comp_row, column=1).border = thin_border

            data_cell(ws7, comp_row, 2, round(r.metrics.mae, 6), True, "0.000000")
            data_cell(ws7, comp_row, 3, round(r.metrics.rmse, 6), True, "0.000000")
            c = data_cell(ws7, comp_row, 4, round(da_pct, 1), True, '0.0"%"')
            if da_pct >= 55:
                c.fill = good_fill
            elif da_pct < 50:
                c.fill = bad_fill

            if naive_result and r.model.value != "Naive":
                rmse_diff = ((r.metrics.rmse - naive_result.metrics.rmse) / naive_result.metrics.rmse * 100) if naive_result.metrics.rmse > 0 else 0
                c = ws7.cell(row=comp_row, column=5, value=f"{rmse_diff:+.2f}%")
                c.font = Font(name="Consolas", size=10, color=GOOD_GREEN if rmse_diff < 0 else BAD_RED)
                c.alignment = center_align
                c.border = thin_border

                mae_diff = ((r.metrics.mae - naive_result.metrics.mae) / naive_result.metrics.mae * 100) if naive_result.metrics.mae > 0 else 0
                c = ws7.cell(row=comp_row, column=6, value=f"{mae_diff:+.2f}%")
                c.font = Font(name="Consolas", size=10, color=GOOD_GREEN if mae_diff < 0 else BAD_RED)
                c.alignment = center_align
                c.border = thin_border

                da_diff = (r.metrics.directionalAccuracy - naive_result.metrics.directionalAccuracy) * 100
                c = ws7.cell(row=comp_row, column=7, value=f"{da_diff:+.1f}pp")
                c.font = Font(name="Consolas", size=10,
                              color=GOOD_GREEN if da_diff > 0 else BAD_RED if da_diff < 0 else "64748B")
                c.alignment = center_align
                c.border = thin_border

                # Verdict
                beats_both = r.metrics.rmse < naive_result.metrics.rmse and r.metrics.directionalAccuracy > naive_result.metrics.directionalAccuracy
                if beats_both:
                    verdict = t["outperforms"]
                    is_outperf, is_partial = True, False
                elif rmse_diff < 0 or da_diff > 0:
                    verdict = t["partial_result"]
                    is_outperf, is_partial = False, True
                else:
                    verdict = t["underperforms"]
                    is_outperf, is_partial = False, False
                c = ws7.cell(row=comp_row, column=8, value=verdict)
                c.font = Font(name="Calibri", bold=True, size=10,
                              color=GOOD_GREEN if is_outperf else WARN_AMBER if is_partial else BAD_RED)
                c.alignment = center_align
                c.border = thin_border
            else:
                for cn in range(5, 9):
                    c = ws7.cell(row=comp_row, column=cn, value=t["baseline_label"])
                    c.font = desc_font
                    c.alignment = center_align
                    c.border = thin_border
            comp_row += 1

        # ── Chart data for backtest ──
        chart_row = comp_row + 2
        chart_col = 10
        n_models = len(sorted_results)

        ws7.cell(row=chart_row - 1, column=chart_col, value=t["model_col"]).font = header_font
        ws7.cell(row=chart_row - 1, column=chart_col + 1, value="RMSE").font = header_font
        ws7.cell(row=chart_row - 1, column=chart_col + 2, value="MAE").font = header_font
        ws7.cell(row=chart_row - 1, column=chart_col + 3, value="DA %").font = header_font
        for i, r in enumerate(sorted_results):
            ws7.cell(row=chart_row + i, column=chart_col, value=r.model.value)
            ws7.cell(row=chart_row + i, column=chart_col + 1, value=round(r.metrics.rmse, 6))
            ws7.cell(row=chart_row + i, column=chart_col + 2, value=round(r.metrics.mae, 6))
            ws7.cell(row=chart_row + i, column=chart_col + 3, value=round(r.metrics.directionalAccuracy * 100, 1))

        # RMSE + MAE chart
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = t["chart_mae_rmse"]
        chart1.y_axis.title = "Error"
        chart1.width = 26
        chart1.height = 14
        d1 = Reference(ws7, min_col=chart_col + 1, max_col=chart_col + 2,
                        min_row=chart_row - 1, max_row=chart_row - 1 + n_models)
        c1 = Reference(ws7, min_col=chart_col, min_row=chart_row, max_row=chart_row - 1 + n_models)
        chart1.add_data(d1, titles_from_data=True)
        chart1.set_categories(c1)
        chart1.series[0].graphicalProperties.solidFill = "3B82F6"
        if len(chart1.series) > 1:
            chart1.series[1].graphicalProperties.solidFill = BRAND_BLUE
        ws7.add_chart(chart1, f"A{comp_row + 2}")

        # DA chart
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = t["chart_da_pct"]
        chart2.y_axis.title = "%"
        chart2.y_axis.scaling.min = 0
        chart2.y_axis.scaling.max = 100
        chart2.width = 26
        chart2.height = 14
        d2 = Reference(ws7, min_col=chart_col + 3, min_row=chart_row - 1, max_row=chart_row - 1 + n_models)
        c2 = Reference(ws7, min_col=chart_col, min_row=chart_row, max_row=chart_row - 1 + n_models)
        chart2.add_data(d2, titles_from_data=True)
        chart2.set_categories(c2)
        for idx_dp, r in enumerate(sorted_results):
            pt = DataPoint(idx=idx_dp)
            da_val = r.metrics.directionalAccuracy * 100
            if da_val >= 55:
                pt.graphicalProperties.solidFill = GOOD_GREEN
            elif da_val >= 50:
                pt.graphicalProperties.solidFill = WARN_AMBER
            else:
                pt.graphicalProperties.solidFill = BAD_RED
            chart2.series[0].data_points.append(pt)
        chart2.legend = None
        ws7.add_chart(chart2, f"A{comp_row + 18}")

        for col_i in range(1, 9):
            ws7.column_dimensions[get_column_letter(col_i)].width = 18

        # ── Sheet 8: WALK-FORWARD WINDOWS ──
        ws8 = wb.create_sheet(t["windows_sheet"])
        ws8.sheet_properties.tabColor = "EF4444"

        ws8.merge_cells("A1:I1")
        ws8["A1"] = f"{t['windows_sheet']}"
        ws8["A1"].font = brand_font
        ws8["A1"].alignment = center_align
        ws8.row_dimensions[1].height = 40

        win_headers = [t["model_col"], t["window_num"], t["train_start"], t["train_end"],
                       t["test_start"], t["test_end"], t["mae_col"], t["rmse_col"], t["da_col"]]
        for ci, h in enumerate(win_headers, 1):
            ws8.cell(row=3, column=ci, value=h)
        style_header_row(ws8, 3, len(win_headers))

        win_row = 4
        for r in sorted_results:
            for wi, w in enumerate(r.windows):
                ws8.cell(row=win_row, column=1, value=r.model.value).font = Font(name="Calibri", bold=True, size=10)
                ws8.cell(row=win_row, column=1).border = thin_border
                ws8.cell(row=win_row, column=2, value=wi + 1).border = thin_border
                ws8.cell(row=win_row, column=2).alignment = center_align
                for ci, val in [(3, w.trainStart), (4, w.trainEnd), (5, w.testStart), (6, w.testEnd)]:
                    ws8.cell(row=win_row, column=ci, value=val).font = Font(name="Calibri", size=10)
                    ws8.cell(row=win_row, column=ci).border = thin_border
                data_cell(ws8, win_row, 7, round(w.mae, 6), True, "0.000000")
                data_cell(ws8, win_row, 8, round(w.rmse, 6), True, "0.000000")
                c = data_cell(ws8, win_row, 9, round(w.directionalAccuracy * 100, 1), True, '0.0"%"')
                da_val = w.directionalAccuracy * 100
                if da_val >= 55:
                    c.fill = good_fill
                elif da_val < 50:
                    c.fill = bad_fill
                win_row += 1

        # Per-model summary
        summ_row = win_row + 2
        ws8.merge_cells(start_row=summ_row, start_column=1, end_row=summ_row, end_column=12)
        ws8.cell(row=summ_row, column=1, value=t["per_model_stats"]).font = section_font
        summ_row += 1

        sm_headers = [t["model_col"], "Windows", "MAE (mean)", "MAE (std)", "MAE (min)", "MAE (max)",
                      "RMSE (mean)", "RMSE (std)", "DA % (mean)", "DA % (std)", "DA % (min)", "DA % (max)"]
        for ci, h in enumerate(sm_headers, 1):
            ws8.cell(row=summ_row, column=ci, value=h)
        style_header_row(ws8, summ_row, len(sm_headers))
        summ_row += 1

        for r in sorted_results:
            if not r.windows:
                continue
            maes = [w.mae for w in r.windows]
            rmses = [w.rmse for w in r.windows]
            das = [w.directionalAccuracy * 100 for w in r.windows]
            ws8.cell(row=summ_row, column=1, value=r.model.value).font = Font(name="Calibri", bold=True, size=10)
            ws8.cell(row=summ_row, column=1).border = thin_border
            data_cell(ws8, summ_row, 2, len(r.windows), True)
            for ci, vals in [(3, maes), (7, rmses)]:
                for j, fn in enumerate([np.mean, np.std, np.min, np.max]):
                    data_cell(ws8, summ_row, ci + j, round(float(fn(vals)), 6), True, "0.000000")
            for j, fn in enumerate([np.mean, np.std, np.min, np.max]):
                c = data_cell(ws8, summ_row, 9 + j, round(float(fn(das)), 1), True, "0.0")
                if j == 0:
                    val_mean = float(np.mean(das))
                    if val_mean >= 55:
                        c.fill = good_fill
                    elif val_mean < 50:
                        c.fill = bad_fill
            summ_row += 1

        # Rolling charts per model
        chart_start = summ_row + 2
        for r in sorted_results:
            if not r.windows:
                continue
            ws8.cell(row=chart_start, column=1, value=f"{r.model.value} — Rolling Performance").font = subtitle_font
            chart_start += 1
            ws8.cell(row=chart_start, column=1, value="Window")
            ws8.cell(row=chart_start, column=2, value="MAE")
            ws8.cell(row=chart_start, column=3, value="RMSE")
            ws8.cell(row=chart_start, column=4, value="DA %")
            for wi, w in enumerate(r.windows):
                ws8.cell(row=chart_start + 1 + wi, column=1, value=f"W{wi + 1}")
                ws8.cell(row=chart_start + 1 + wi, column=2, value=round(w.mae, 6))
                ws8.cell(row=chart_start + 1 + wi, column=3, value=round(w.rmse, 6))
                ws8.cell(row=chart_start + 1 + wi, column=4, value=round(w.directionalAccuracy * 100, 1))

            lc = LineChart()
            lc.title = f"{r.model.value} — MAE & RMSE across windows"
            lc.style = 10
            lc.width = 28
            lc.height = 12
            d = Reference(ws8, min_col=2, max_col=3, min_row=chart_start, max_row=chart_start + len(r.windows))
            c_ref = Reference(ws8, min_col=1, min_row=chart_start + 1, max_row=chart_start + len(r.windows))
            lc.add_data(d, titles_from_data=True)
            lc.set_categories(c_ref)
            lc.series[0].graphicalProperties.line.solidFill = "3B82F6"
            if len(lc.series) > 1:
                lc.series[1].graphicalProperties.line.solidFill = BRAND_BLUE
            ws8.add_chart(lc, f"F{chart_start}")
            chart_start += len(r.windows) + 18

        for col_i in range(1, 13):
            ws8.column_dimensions[get_column_letter(col_i)].width = 16

        # ── Sheet 9: ERROR ANALYSIS (NEW) ──
        ws9 = wb.create_sheet(t["error_sheet"])
        ws9.sheet_properties.tabColor = "EC4899"

        ws9.merge_cells("A1:H1")
        ws9["A1"] = f"{t['error_sheet']}"
        ws9["A1"].font = brand_font
        ws9["A1"].alignment = center_align
        ws9.row_dimensions[1].height = 40

        row = 3
        row = write_section_title(ws9, row, 1, t["error_dist_per_model"], end_col=8)

        err_headers = [t["model_col"], t["mean_error"], t["std_error"], t["min_error"], t["max_error"],
                       t["median_error"], t["error_skewness"], t["error_kurtosis"]]
        for ci, h in enumerate(err_headers, 1):
            ws9.cell(row=row, column=ci, value=h)
        style_header_row(ws9, row, len(err_headers))
        row += 1

        for r in sorted_results:
            w_maes = [w.mae for w in r.windows]
            w_rmses = [w.rmse for w in r.windows]
            if len(w_maes) > 1:
                err_arr = np.array(w_maes)
                ws9.cell(row=row, column=1, value=r.model.value).font = Font(name="Calibri", bold=True, size=10)
                ws9.cell(row=row, column=1).border = thin_border
                data_cell(ws9, row, 2, round(float(err_arr.mean()), 6), True, "0.000000")
                data_cell(ws9, row, 3, round(float(err_arr.std()), 6), True, "0.000000")
                data_cell(ws9, row, 4, round(float(err_arr.min()), 6), True, "0.000000")
                data_cell(ws9, row, 5, round(float(err_arr.max()), 6), True, "0.000000")
                data_cell(ws9, row, 6, round(float(np.median(err_arr)), 6), True, "0.000000")
                from scipy import stats as sp_stats
                try:
                    data_cell(ws9, row, 7, round(float(sp_stats.skew(err_arr)), 4), True, "0.0000")
                    data_cell(ws9, row, 8, round(float(sp_stats.kurtosis(err_arr)), 4), True, "0.0000")
                except Exception:
                    pass
                row += 1

        # Error percentile table
        row += 1
        row = write_section_title(ws9, row, 1, t["mae_percentiles"], end_col=8)

        pct_headers = [t["model_col"], "P10", "P25", "P50", "P75", "P90", "P95", "IQR"]
        for ci, h in enumerate(pct_headers, 1):
            ws9.cell(row=row, column=ci, value=h)
        style_header_row(ws9, row, len(pct_headers))
        row += 1

        for r in sorted_results:
            w_maes = np.array([w.mae for w in r.windows])
            if len(w_maes) > 2:
                ws9.cell(row=row, column=1, value=r.model.value).font = Font(name="Calibri", bold=True, size=10)
                ws9.cell(row=row, column=1).border = thin_border
                for ci, pct in enumerate([10, 25, 50, 75, 90, 95], 2):
                    data_cell(ws9, row, ci, round(float(np.percentile(w_maes, pct)), 6), True, "0.000000")
                iqr = float(np.percentile(w_maes, 75) - np.percentile(w_maes, 25))
                data_cell(ws9, row, 8, round(iqr, 6), True, "0.000000")
                row += 1

        # Consistency Score
        row += 1
        row = write_section_title(ws9, row, 1, t["consistency_score"], end_col=8)
        ws9.cell(row=row, column=1, value=t["consistency_desc"]).font = desc_font
        ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        cons_headers = [t["model_col"], "MAE CV %", "RMSE CV %", "DA CV %", t["consistency_score"]]
        for ci, h in enumerate(cons_headers, 1):
            ws9.cell(row=row, column=ci, value=h)
        style_header_row(ws9, row, len(cons_headers))
        row += 1

        for r in sorted_results:
            maes_arr = np.array([w.mae for w in r.windows])
            rmses_arr = np.array([w.rmse for w in r.windows])
            das_arr = np.array([w.directionalAccuracy * 100 for w in r.windows])
            if len(maes_arr) > 1 and np.mean(maes_arr) > 0:
                mae_cv = float(np.std(maes_arr) / np.mean(maes_arr) * 100)
                rmse_cv = float(np.std(rmses_arr) / np.mean(rmses_arr) * 100)
                da_cv = float(np.std(das_arr) / np.mean(das_arr) * 100) if np.mean(das_arr) > 0 else 0

                ws9.cell(row=row, column=1, value=r.model.value).font = Font(name="Calibri", bold=True, size=10)
                ws9.cell(row=row, column=1).border = thin_border
                data_cell(ws9, row, 2, round(mae_cv, 2), True, "0.00")
                data_cell(ws9, row, 3, round(rmse_cv, 2), True, "0.00")
                data_cell(ws9, row, 4, round(da_cv, 2), True, "0.00")

                avg_cv = (mae_cv + rmse_cv) / 2
                if avg_cv < 15:
                    rating = t["cons_excellent"]
                    r_color = GOOD_GREEN
                elif avg_cv < 30:
                    rating = t["cons_good"]
                    r_color = ACCENT_TEAL
                elif avg_cv < 50:
                    rating = t["cons_moderate"]
                    r_color = WARN_AMBER
                else:
                    rating = t["cons_poor"]
                    r_color = BAD_RED

                c = ws9.cell(row=row, column=5, value=rating)
                c.font = Font(name="Calibri", bold=True, size=11, color=r_color)
                c.alignment = center_align
                c.border = thin_border
                row += 1

        # ── Top-N biggest errors per model (by window MAE) ──
        row += 1
        row = write_section_title(ws9, row, 1, t["top_errors_title"], end_col=8)

        top_n = 5
        for r in sorted_results:
            if not r.windows or len(r.windows) < 2:
                continue
            # Sort windows by MAE descending
            sorted_wins = sorted(enumerate(r.windows), key=lambda x: x[1].mae, reverse=True)[:top_n]
            ws9.cell(row=row, column=1, value=r.model.value).font = Font(name="Calibri", bold=True, size=10, color=BRAND_BLUE)
            ws9.cell(row=row, column=1).border = thin_border
            row += 1
            ws9.cell(row=row, column=1, value=t["window_label"]).font = header_font
            ws9.cell(row=row, column=1).fill = header_fill
            ws9.cell(row=row, column=1).border = thin_border
            ws9.cell(row=row, column=2, value="MAE").font = header_font
            ws9.cell(row=row, column=2).fill = header_fill
            ws9.cell(row=row, column=2).border = thin_border
            ws9.cell(row=row, column=3, value="RMSE").font = header_font
            ws9.cell(row=row, column=3).fill = header_fill
            ws9.cell(row=row, column=3).border = thin_border
            ws9.cell(row=row, column=4, value="DA %").font = header_font
            ws9.cell(row=row, column=4).fill = header_fill
            ws9.cell(row=row, column=4).border = thin_border
            ws9.cell(row=row, column=5, value=t["test_start"]).font = header_font
            ws9.cell(row=row, column=5).fill = header_fill
            ws9.cell(row=row, column=5).border = thin_border
            row += 1
            for wi, w in sorted_wins:
                data_cell(ws9, row, 1, f"W{wi + 1}", False)
                data_cell(ws9, row, 2, round(w.mae, 6), True, "0.000000")
                data_cell(ws9, row, 3, round(w.rmse, 6), True, "0.000000")
                data_cell(ws9, row, 4, round(w.directionalAccuracy * 100, 1), True, "0.0")
                data_cell(ws9, row, 5, w.testStart if hasattr(w, 'testStart') else "", False)
                row += 1
            row += 1

        for col_i in range(1, 9):
            ws9.column_dimensions[get_column_letter(col_i)].width = 18

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 10: STATISTICAL TESTS (enhanced)
    # ═══════════════════════════════════════════════════════════════════════════
    if tests_data and req.includeTests:
        ws10 = wb.create_sheet(t["tests_sheet"])
        ws10.sheet_properties.tabColor = "10B981"

        ws10.merge_cells("A1:E1")
        ws10["A1"] = f"{t['tests_sheet']}"
        ws10["A1"].font = brand_font
        ws10["A1"].alignment = center_align
        ws10.row_dimensions[1].height = 40

        row = 3
        # ADF Test
        row = write_section_title(ws10, row, 1, t["adf_title"], end_col=5)
        ws10.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws10.cell(row=row, column=1, value=t["adf_desc"]).font = desc_font
        ws10.cell(row=row, column=1).alignment = wrap_align
        row += 1

        for label, val in [
            (t["statistic"], round(tests_data.adf.statistic, 6)),
            (t["pvalue"], _format_pval(tests_data.adf.pValue)),
            (t["result"], t["stationary"] if tests_data.adf.isStationary else t["not_stationary"]),
        ]:
            ws10.cell(row=row, column=1, value=label).font = subtitle_font
            ws10.cell(row=row, column=1).border = thin_border
            ws10.cell(row=row, column=1).fill = light_fill
            c = ws10.cell(row=row, column=2, value=val)
            c.font = mono_font if isinstance(val, (int, float)) else normal_font
            c.border = thin_border
            if label == t["result"]:
                c.fill = good_fill if tests_data.adf.isStationary else bad_fill
                c.font = Font(name="Calibri", bold=True, size=11,
                              color=GOOD_GREEN if tests_data.adf.isStationary else BAD_RED)
            row += 1

        row += 1
        # Ljung-Box
        row = write_section_title(ws10, row, 1, t["lb_title"], end_col=5)
        ws10.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws10.cell(row=row, column=1, value=t["lb_desc"]).font = desc_font
        ws10.cell(row=row, column=1).alignment = wrap_align
        row += 1

        # Add Ljung-Box specifics
        lb_best_model = sorted_results[overall_best_idx].model.value if sorted_results else "N/A"
        for label, val in [
            (t["lb_residuals"], t["lb_residuals_of"].format(lb_best_model)),
            (t["lb_lag"], t["lb_lag_default"]),
            (t["statistic"], round(tests_data.ljungBox.statistic, 6)),
            (t["pvalue"], _format_pval(tests_data.ljungBox.pValue)),
            (t["result"], t["no_autocorr"] if tests_data.ljungBox.noAutocorrelation else t["autocorr_detected"]),
        ]:
            ws10.cell(row=row, column=1, value=label).font = subtitle_font
            ws10.cell(row=row, column=1).border = thin_border
            ws10.cell(row=row, column=1).fill = light_fill
            c = ws10.cell(row=row, column=2, value=val)
            c.font = mono_font if isinstance(val, (int, float)) else normal_font
            c.border = thin_border
            if label == t["result"]:
                c.fill = good_fill if tests_data.ljungBox.noAutocorrelation else warn_fill
                c.font = Font(name="Calibri", bold=True, size=11,
                              color=GOOD_GREEN if tests_data.ljungBox.noAutocorrelation else WARN_AMBER)
            row += 1

        # Diebold-Mariano
        if tests_data.dieboldMariano:
            row += 1
            row = write_section_title(ws10, row, 1, t["dm_title"], end_col=5)
            ws10.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws10.cell(row=row, column=1, value=t["dm_desc"]).font = desc_font
            ws10.cell(row=row, column=1).alignment = wrap_align
            row += 1

            dm = tests_data.dieboldMariano
            dm_sig = dm.pValue < 0.05
            for label, val in [
                (t["statistic"], round(dm.statistic, 6)),
                (t["pvalue"], _format_pval(dm.pValue)),
                (t["result"], t["significant"] if dm_sig else t["not_significant"]),
                (t["better_model"], dm.betterModel.value if dm.betterModel else "N/A"),
            ]:
                ws10.cell(row=row, column=1, value=label).font = subtitle_font
                ws10.cell(row=row, column=1).border = thin_border
                ws10.cell(row=row, column=1).fill = light_fill
                c = ws10.cell(row=row, column=2, value=val)
                c.font = mono_font if isinstance(val, (int, float)) else normal_font
                c.border = thin_border
                if label == t["result"]:
                    c.fill = good_fill if dm_sig else light_fill
                row += 1

        # Significance summary
        row += 2
        row = write_section_title(ws10, row, 1, t["significance_summary"], end_col=5)

        sig_items = []
        if tests_data.adf.isStationary:
            sig_items.append(("ADF [OK]", t["sig_adf_ok"]))
        else:
            sig_items.append(("ADF [!]", t["sig_adf_fail"]))

        if tests_data.ljungBox.noAutocorrelation:
            sig_items.append(("Ljung-Box [OK]", t["sig_lb_ok"]))
        else:
            sig_items.append(("Ljung-Box [!]", t["sig_lb_fail"]))

        if tests_data.dieboldMariano:
            dm = tests_data.dieboldMariano
            if dm.pValue < 0.05:
                better = dm.betterModel.value if dm.betterModel else "N/A"
                sig_items.append(("DM [OK]", t["sig_dm_ok"].format(better)))
            else:
                sig_items.append(("DM [!]", t["sig_dm_fail"]))

        for test_name, interp in sig_items:
            ws10.cell(row=row, column=1, value=test_name).font = Font(name="Calibri", bold=True, size=10)
            ws10.cell(row=row, column=1).border = thin_border
            ws10.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            ws10.cell(row=row, column=2, value=interp).font = desc_font
            ws10.cell(row=row, column=2).alignment = wrap_align
            ws10.cell(row=row, column=2).border = thin_border
            row += 1

        ws10.column_dimensions["A"].width = 25
        ws10.column_dimensions["B"].width = 50
        ws10.column_dimensions["C"].width = 20
        ws10.column_dimensions["D"].width = 20
        ws10.column_dimensions["E"].width = 20

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 11: CONCLUSION (enhanced)
    # ═══════════════════════════════════════════════════════════════════════════
    if backtest_data:
        ws_conc = wb.create_sheet(t["conclusion_sheet"])
        ws_conc.sheet_properties.tabColor = "6366F1"

        ws_conc.merge_cells("A1:F1")
        ws_conc["A1"] = f"{t['conclusion_sheet']}"
        ws_conc["A1"].font = brand_font
        ws_conc["A1"].alignment = center_align
        ws_conc.row_dimensions[1].height = 40

        c_row = 3
        c_row = write_section_title(ws_conc, c_row, 1, t["key_findings"])

        overall_best = sorted_results[overall_best_idx]
        ob_da = overall_best.metrics.directionalAccuracy * 100

        findings = [
            t["finding_best_model"].format(overall_best.model.value, f"{overall_best.metrics.rmse:.6f}", f"{ob_da:.1f}"),
            t["finding_model_count"].format(len(sorted_results)),
            t["finding_data_points"].format(len(close)),
            t["finding_period"].format(data_start, data_end),
            t["finding_naive_da"].format(f"{naive_da * 100:.1f}"),
            t["finding_ann_vol"].format(f"{ann_vol * 100:.2f}"),
            t["finding_max_dd"].format(f"{max_dd * 100:.2f}"),
            t["finding_total_change"].format(f"{total_change_pct:.3f}"),
        ]
        if genuine_winners:
            findings.append(t["finding_winners"].format(", ".join(genuine_winners)))
        else:
            findings.append(t["finding_no_winners"])

        if ob_da < 50:
            findings.append(t["finding_da_below_50"].format(f"{ob_da:.1f}"))
        elif ob_da < 55:
            findings.append(t["finding_da_near_50"].format(f"{ob_da:.1f}"))

        for finding in findings:
            ws_conc.cell(row=c_row, column=1, value="•").font = normal_font
            ws_conc.merge_cells(start_row=c_row, start_column=2, end_row=c_row, end_column=6)
            ws_conc.cell(row=c_row, column=2, value=finding).font = normal_font
            ws_conc.cell(row=c_row, column=2).alignment = wrap_align
            c_row += 1

        # Final ranking
        c_row += 1
        c_row = write_section_title(ws_conc, c_row, 1, t["final_ranking"])

        rank_headers = ["#", t["model_col"], "RMSE", "DA %", t["beats_naive_col"], t["verdict"]]
        for ci, h in enumerate(rank_headers, 1):
            ws_conc.cell(row=c_row, column=ci, value=h)
        style_header_row(ws_conc, c_row, len(rank_headers))
        c_row += 1

        def model_sort_key(r):
            beats = (r.metrics.rmse < naive_result.metrics.rmse and
                     r.metrics.directionalAccuracy > naive_result.metrics.directionalAccuracy) if naive_result else False
            return (0 if beats else 1, r.metrics.rmse)

        final_ranked = sorted(sorted_results, key=model_sort_key)
        for rank_i, r in enumerate(final_ranked, 1):
            da_pct = r.metrics.directionalAccuracy * 100
            beats = False
            if naive_result and r.model.value != "Naive":
                beats = (r.metrics.rmse < naive_result.metrics.rmse and
                        r.metrics.directionalAccuracy > naive_result.metrics.directionalAccuracy)

            c = data_cell(ws_conc, c_row, 1, rank_i, True)
            c.alignment = center_align

            ws_conc.cell(row=c_row, column=2, value=r.model.value).font = Font(
                name="Calibri", bold=True, size=11,
                color=BRAND_BLUE if r.model.value == overall_best.model.value else BRAND_DARK)
            ws_conc.cell(row=c_row, column=2).border = thin_border

            data_cell(ws_conc, c_row, 3, round(r.metrics.rmse, 6), True, "0.000000")
            c = data_cell(ws_conc, c_row, 4, round(da_pct, 1), True, "0.0")
            if da_pct >= 55:
                c.fill = good_fill
            elif da_pct < 50:
                c.fill = bad_fill

            beats_text = t.get("yes_label", "Ano") if beats else (t.get("baseline_label", "zaklad") if r.model.value == "Naive" else t.get("no_label", "Nie"))
            c = ws_conc.cell(row=c_row, column=5, value=beats_text)
            c.font = normal_font
            c.alignment = center_align
            c.border = thin_border

            if r.model.value == overall_best.model.value:
                verdict_text = t.get("best_label", "NAJLEPSI")
                v_color = GOOD_GREEN
            elif beats:
                verdict_text = t.get("strong_label", "Silny")
                v_color = ACCENT_TEAL
            elif da_pct >= 50:
                verdict_text = t.get("moderate_label", "Priemerny")
                v_color = WARN_AMBER
            else:
                verdict_text = t.get("weak_label", "Slaby")
                v_color = BAD_RED
            c = ws_conc.cell(row=c_row, column=6, value=verdict_text)
            c.font = Font(name="Calibri", bold=True, size=10, color=v_color)
            c.alignment = center_align
            c.border = thin_border
            c_row += 1

        # ── Statistical tests conclusion ──
        if tests_data:
            c_row += 1
            c_row = write_section_title(ws_conc, c_row, 1, t["stat_tests_conclusion"])

            stat_conclusions = []
            adf_p = tests_data.adf.pValue
            adf_state = t["conc_adf_stationary"] if tests_data.adf.isStationary else t["conc_adf_nonstationary"]
            adf_comment = t["conc_adf_ok_comment"] if tests_data.adf.isStationary else t["conc_adf_fail_comment"]
            stat_conclusions.append(f"ADF test: p = {_format_pval(adf_p)} → {adf_state}. {adf_comment}")

            lb_p = tests_data.ljungBox.pValue
            lb_result = t["conc_lb_ok"] if tests_data.ljungBox.noAutocorrelation else t["conc_lb_fail"]
            stat_conclusions.append(f"Ljung-Box test: p = {_format_pval(lb_p)} → {lb_result}.")

            if tests_data.dieboldMariano:
                dm = tests_data.dieboldMariano
                dm_better = dm.betterModel.value if dm.betterModel else "N/A"
                dm_result = t["conc_dm_ok"].format(dm_better) if dm.pValue < 0.05 else t["conc_dm_fail"]
                stat_conclusions.append(f"Diebold-Mariano test: p = {_format_pval(dm.pValue)} → {dm_result}.")

            for sc in stat_conclusions:
                ws_conc.cell(row=c_row, column=1, value="•").font = normal_font
                ws_conc.merge_cells(start_row=c_row, start_column=2, end_row=c_row, end_column=6)
                ws_conc.cell(row=c_row, column=2, value=sc).font = desc_font
                ws_conc.cell(row=c_row, column=2).alignment = wrap_align
                c_row += 1

        # ── Methodology note ──
        c_row += 1
        c_row = write_section_title(ws_conc, c_row, 1, t["methodology_title"])
        methods = [
            t["method_walkforward"],
            t["method_windows"].format(max(30, len(close) // 3), max(5, len(close) // 10)),
            t["method_metrics"],
            t["method_tests"],
            t["method_datasource"],
            t["method_generated"].format(now.strftime('%Y-%m-%d %H:%M UTC')),
        ]
        for method in methods:
            ws_conc.cell(row=c_row, column=1, value="•").font = normal_font
            ws_conc.merge_cells(start_row=c_row, start_column=2, end_row=c_row, end_column=6)
            ws_conc.cell(row=c_row, column=2, value=method).font = desc_font
            ws_conc.cell(row=c_row, column=2).alignment = wrap_align
            c_row += 1

        # ── Limitations ──
        c_row += 1
        c_row = write_section_title(ws_conc, c_row, 1, t["limitations_title"])
        limitations = [
            t["lim_timeframe"].format(req.timeframe.value, symbol),
            t["lim_interval"].format(data_start, data_end, len(close)),
            t["lim_no_exogenous"],
            t["lim_yahoo"],
            t["lim_ai_api"],
        ]
        for lim in limitations:
            ws_conc.cell(row=c_row, column=1, value="•").font = normal_font
            ws_conc.merge_cells(start_row=c_row, start_column=2, end_row=c_row, end_column=6)
            ws_conc.cell(row=c_row, column=2, value=lim).font = desc_font
            ws_conc.cell(row=c_row, column=2).alignment = wrap_align
            c_row += 1

        # ── Improvements ──
        c_row += 1
        c_row = write_section_title(ws_conc, c_row, 1, t["improvements_title"])
        improvements = [
            t["imp_more_pairs"],
            t["imp_more_features"],
            t["imp_deep_models"],
            t["imp_validation"],
            t["imp_exogenous"],
            t["imp_oos_verify"],
        ]
        for imp in improvements:
            ws_conc.cell(row=c_row, column=1, value="•").font = normal_font
            ws_conc.merge_cells(start_row=c_row, start_column=2, end_row=c_row, end_column=6)
            ws_conc.cell(row=c_row, column=2, value=imp).font = desc_font
            ws_conc.cell(row=c_row, column=2).alignment = wrap_align
            c_row += 1

        ws_conc.column_dimensions["A"].width = 5
        ws_conc.column_dimensions["B"].width = 22
        ws_conc.column_dimensions["C"].width = 16
        ws_conc.column_dimensions["D"].width = 12
        ws_conc.column_dimensions["E"].width = 18
        ws_conc.column_dimensions["F"].width = 15

    # ═══════════════════════════════════════════════════════════════════════════
    # FREEZE PANES — lock headers on sheets with large tables
    # ═══════════════════════════════════════════════════════════════════════════
    # Price Data (ws3): freeze row 1 (header)
    try:
        ws3.freeze_panes = "A2"
    except Exception:
        pass

    # Technical Indicators (ws4): freeze row 3 (header)
    try:
        ws4.freeze_panes = "A4"
    except Exception:
        pass

    # Walk-Forward Windows (ws8) if it exists
    for _ws_name in wb.sheetnames:
        _ws = wb[_ws_name]
        if _ws_name == t["windows_sheet"]:
            try:
                _ws.freeze_panes = "A4"
            except Exception:
                pass
        if _ws_name == t["backtest_sheet"]:
            try:
                _ws.freeze_panes = "A7"
            except Exception:
                pass

    # ═══════════════════════════════════════════════════════════════════════════
    # Save
    # ═══════════════════════════════════════════════════════════════════════════
    report_path = settings.reports_path / f"{report_id}.xlsx"
    wb.save(str(report_path))

    log.info("excel_report.generated", report_id=report_id, path=str(report_path),
             sheets=len(wb.sheetnames))
    return report_id, report_path
