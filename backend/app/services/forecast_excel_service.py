"""Excel report generation for individual model forecasts.

Produces a detailed .xlsx with:
- Model training info (type, params, target, metrics)
- Historical price data used for training
- Forecast predictions with dates and confidence intervals
- Charts: historical prices + forecast line
- Summary section for "check tomorrow" comparison
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.logging_config import get_logger
from app.models.orm import TrainedModelRow
from app.schemas.schemas import ForecastPoint
from app.services.model_service import forecast
from app.services.ohlc_service import get_close_series, get_ohlc_dataframe

log = get_logger(__name__)

# ─── Colors ────────────────────────────────────────────────────────────────────

BRAND_BLUE = "0EA5E9"
BRAND_DARK = "1E293B"
HEADER_BG = "0F172A"
HEADER_FG = "FFFFFF"
GOOD_GREEN = "16A34A"
BAD_RED = "DC2626"
LIGHT_BG = "F1F5F9"
BORDER_COLOR = "CBD5E1"
FORECAST_COLOR = "F59E0B"  # amber for forecast


async def generate_forecast_excel(
    db: AsyncSession,
    model_id: str,
    horizon: int,
) -> tuple[str, Path, str]:
    """Generate forecast Excel report. Returns (report_id, file_path, filename)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    # ── Load model info ────────────────────────────────────────────────────────
    stmt = select(TrainedModelRow).where(TrainedModelRow.id == model_id)
    result = await db.execute(stmt)
    model_row = result.scalar_one_or_none()
    if model_row is None:
        raise ValueError(f"Model {model_id} not found")

    symbol = model_row.symbol
    timeframe = model_row.timeframe
    model_type = model_row.model_type
    target_type = model_row.target_type or "close"
    config = model_row.config_json or {}
    features_cfg = config.get("features", {})
    hyperparams = config.get("hyperparams", {})
    metrics = model_row.metrics_json or {}
    trained_at = model_row.trained_at

    # ── Generate forecast ──────────────────────────────────────────────────────
    forecast_points = await forecast(db, model_id, horizon)

    # ── Load historical OHLC ───────────────────────────────────────────────────
    import joblib
    if not model_row.artifact_path:
        raise ValueError(f"Model {model_id} has no saved artifact path")
    artifact = joblib.load(model_row.artifact_path)
    last_values = artifact.get("last_values", [])
    last_index = artifact.get("last_index", [])

    # Try to get richer OHLC data from DB
    ohlc_df = None
    if last_index:
        start_date = last_index[0][:10]
        end_date = last_index[-1][:10]
        ohlc_df = await get_ohlc_dataframe(db, symbol, timeframe, start_date, end_date)

    report_id = str(uuid.uuid4())[:8]
    now = datetime.now(timezone.utc)

    # ── Styles ──────────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    brand_font = Font(name="Calibri", bold=True, color=BRAND_BLUE, size=14)
    title_font = Font(name="Calibri", bold=True, size=13, color=BRAND_DARK)
    subtitle_font = Font(name="Calibri", bold=True, size=11, color=BRAND_DARK)
    normal_font = Font(name="Calibri", size=11)
    mono_font = Font(name="Consolas", size=11)
    desc_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    forecast_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1: Forecast Summary
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Forecast Summary"
    ws.sheet_properties.tabColor = BRAND_BLUE

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"FX Forecast Report — {symbol}"
    ws["A1"].font = brand_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Model: {model_type} | Target: {target_type} | Timeframe: {timeframe} | Horizon: {horizon} bars"
    ws["A2"].font = Font(name="Calibri", size=11, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Model Details ──────────────────────────────────────────────────────────
    row = 4
    ws.cell(row=row, column=1, value="Model Details").font = title_font
    row += 1

    detail_rows = [
        ("Currency Pair", symbol),
        ("Timeframe", timeframe),
        ("Model Type", model_type),
        ("Prediction Target", "Close Price" if target_type == "close" else "Log Return"),
        ("Trained At", trained_at.strftime("%Y-%m-%d %H:%M UTC") if trained_at else "N/A"),
        ("Report Generated", now.strftime("%Y-%m-%d %H:%M UTC")),
        ("Forecast Horizon", f"{horizon} {'trading days' if timeframe == '1D' else 'bars'}"),
    ]

    # Model-specific params
    if model_type == "ARIMA":
        p, d, q = hyperparams.get("p", 1), hyperparams.get("d", 1), hyperparams.get("q", 1)
        auto = hyperparams.get("auto", False)
        detail_rows.append(("ARIMA Order (p,d,q)", f"({p},{d},{q})" + (" [auto-selected]" if auto else "")))
    elif model_type == "MovingAverage":
        detail_rows.append(("Window Size", hyperparams.get("window", 20)))
    elif model_type == "Ridge":
        detail_rows.append(("Alpha (regularization)", hyperparams.get("alpha", 1.0)))
    elif model_type == "RandomForest":
        detail_rows.append(("Number of Trees", hyperparams.get("n_estimators", 100)))
        detail_rows.append(("Max Depth", hyperparams.get("max_depth", 10)))

    # Features (for ML models)
    if features_cfg:
        active_features = []
        if features_cfg.get("lagReturns"):
            active_features.append(f"Lag Returns (n={features_cfg.get('numLags', 5)})")
        if features_cfg.get("sma"):
            active_features.append("SMA(20)")
        if features_cfg.get("ema"):
            active_features.append("EMA(50)")
        if features_cfg.get("rsi"):
            active_features.append("RSI(14)")
        if features_cfg.get("macd"):
            active_features.append("MACD")
        if active_features:
            detail_rows.append(("Features Used", ", ".join(active_features)))

    # Metrics
    if metrics:
        detail_rows.append(("Validation MAE", round(metrics.get("mae", 0), 6)))
        detail_rows.append(("Validation RMSE", round(metrics.get("rmse", 0), 6)))

    # Training data info
    if last_values:
        detail_rows.append(("Training Data Points", len(last_values)))
        detail_rows.append(("Last Training Price", round(last_values[-1], 6)))
        if last_index:
            detail_rows.append(("Training Period", f"{last_index[0][:10]} — {last_index[-1][:10]}"))

    for label, val in detail_rows:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=row, column=1).border = thin_border
        c = ws.cell(row=row, column=2, value=val)
        c.font = mono_font if isinstance(val, (int, float)) else normal_font
        c.border = thin_border
        if isinstance(val, float):
            c.number_format = "0.000000"
        row += 1

    # ── Forecast Key Numbers ───────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="Forecast Key Numbers").font = title_font
    row += 1

    if forecast_points:
        first_pred = forecast_points[0].predicted
        last_pred = forecast_points[-1].predicted
        last_price = last_values[-1] if last_values else first_pred
        total_change = ((last_pred - last_price) / last_price * 100) if last_price else 0
        direction = "UP ↑" if last_pred > last_price else "DOWN ↓" if last_pred < last_price else "FLAT →"

        key_numbers = [
            ("Current Price (last training)", round(last_price, 6)),
            ("First Forecast Value", round(first_pred, 6)),
            ("Last Forecast Value", round(last_pred, 6)),
            ("Total Predicted Change", f"{total_change:+.4f}%"),
            ("Predicted Direction", direction),
            ("Min Forecast Value", round(min(fp.predicted for fp in forecast_points), 6)),
            ("Max Forecast Value", round(max(fp.predicted for fp in forecast_points), 6)),
        ]

        # Next day prediction (most useful for "check tomorrow")
        key_numbers.append(("—" * 20, "—" * 20))
        key_numbers.append(("★ NEXT BAR PREDICTION", ""))
        key_numbers.append(("Date", forecast_points[0].time))
        key_numbers.append(("Predicted Price", round(first_pred, 6)))
        if forecast_points[0].lower is not None:
            key_numbers.append(("Confidence Lower", round(forecast_points[0].lower, 6)))
            key_numbers.append(("Confidence Upper", round(forecast_points[0].upper, 6)))

        for label, val in key_numbers:
            ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=row, column=1).border = thin_border
            c = ws.cell(row=row, column=2, value=val)
            c.border = thin_border
            if isinstance(val, float):
                c.font = mono_font
                c.number_format = "0.000000"
            elif "★" in str(label):
                c.font = Font(name="Calibri", bold=True, size=11, color=BRAND_BLUE)
                ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color=BRAND_BLUE)
            elif "UP" in str(val):
                c.font = Font(name="Calibri", bold=True, size=11, color=GOOD_GREEN)
            elif "DOWN" in str(val):
                c.font = Font(name="Calibri", bold=True, size=11, color=BAD_RED)
            else:
                c.font = normal_font
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2: Forecast Data
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Forecast Data")
    ws2.sheet_properties.tabColor = FORECAST_COLOR

    fc_headers = ["Date", "Predicted Price", "Lower CI", "Upper CI", "Day #"]
    for col, h in enumerate(fc_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(fc_headers))

    for i, fp in enumerate(forecast_points, 2):
        ws2.cell(row=i, column=1, value=fp.time).font = normal_font
        ws2.cell(row=i, column=1).border = thin_border
        c = ws2.cell(row=i, column=2, value=fp.predicted)
        c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
        c.fill = forecast_fill

        if fp.lower is not None:
            c = ws2.cell(row=i, column=3, value=fp.lower)
            c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
        if fp.upper is not None:
            c = ws2.cell(row=i, column=4, value=fp.upper)
            c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border

        ws2.cell(row=i, column=5, value=i - 1).font = mono_font
        ws2.cell(row=i, column=5).alignment = center_align
        ws2.cell(row=i, column=5).border = thin_border

    # Forecast chart
    if len(forecast_points) > 1:
        fc_chart = LineChart()
        fc_chart.title = f"{symbol} — {model_type} Forecast ({horizon} bars)"
        fc_chart.style = 10
        fc_chart.y_axis.title = "Price"
        fc_chart.x_axis.title = "Date"
        fc_chart.width = 30
        fc_chart.height = 15

        n_pts = len(forecast_points)
        pred_ref = Reference(ws2, min_col=2, min_row=1, max_row=n_pts + 1)
        dates_ref = Reference(ws2, min_col=1, min_row=2, max_row=n_pts + 1)
        fc_chart.add_data(pred_ref, titles_from_data=True)
        fc_chart.set_categories(dates_ref)
        fc_chart.series[0].graphicalProperties.line.solidFill = FORECAST_COLOR
        fc_chart.series[0].graphicalProperties.line.width = 25000

        # Add CI bands if available
        if forecast_points[0].lower is not None:
            lower_ref = Reference(ws2, min_col=3, min_row=1, max_row=n_pts + 1)
            upper_ref = Reference(ws2, min_col=4, min_row=1, max_row=n_pts + 1)
            fc_chart.add_data(lower_ref, titles_from_data=True)
            fc_chart.add_data(upper_ref, titles_from_data=True)
            fc_chart.series[1].graphicalProperties.line.solidFill = "94A3B8"
            fc_chart.series[1].graphicalProperties.line.dashStyle = "dash"
            fc_chart.series[1].graphicalProperties.line.width = 12000
            fc_chart.series[2].graphicalProperties.line.solidFill = "94A3B8"
            fc_chart.series[2].graphicalProperties.line.dashStyle = "dash"
            fc_chart.series[2].graphicalProperties.line.width = 12000

        ws2.add_chart(fc_chart, f"G2")

    for i in range(1, 6):
        ws2.column_dimensions[get_column_letter(i)].width = 20

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 3: Historical Price Data
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Training Data")
    ws3.sheet_properties.tabColor = "3B82F6"

    if ohlc_df is not None and len(ohlc_df) > 0:
        hist_headers = ["Date", "Open", "High", "Low", "Close", "Change", "Change %"]
        for col, h in enumerate(hist_headers, 1):
            ws3.cell(row=1, column=col, value=h)
        style_header_row(ws3, 1, len(hist_headers))

        df = ohlc_df.copy()
        if "close" in df.columns:
            df["change"] = df["close"].diff()
            df["change_pct"] = df["close"].pct_change() * 100

        for i, (idx, row_data) in enumerate(df.iterrows(), 2):
            ws3.cell(row=i, column=1, value=str(idx)[:19]).font = normal_font
            ws3.cell(row=i, column=1).border = thin_border
            for j, col_name in enumerate(["open", "high", "low", "close"], 2):
                c = ws3.cell(row=i, column=j, value=round(float(row_data.get(col_name, 0)), 6))
                c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
            if "change" in df.columns and pd.notna(row_data.get("change")):
                c = ws3.cell(row=i, column=6, value=round(float(row_data["change"]), 6))
                c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
                if row_data["change"] > 0:
                    c.font = Font(name="Consolas", size=11, color=GOOD_GREEN)
                elif row_data["change"] < 0:
                    c.font = Font(name="Consolas", size=11, color=BAD_RED)
            if "change_pct" in df.columns and pd.notna(row_data.get("change_pct")):
                c = ws3.cell(row=i, column=7, value=round(float(row_data["change_pct"]), 4))
                c.font = mono_font; c.number_format = '0.0000"%"'; c.border = thin_border

        # Historical price chart
        if len(df) > 1:
            hist_chart = LineChart()
            hist_chart.title = f"{symbol} Historical Prices (Training Data)"
            hist_chart.style = 10
            hist_chart.y_axis.title = "Price"
            hist_chart.width = 30; hist_chart.height = 15
            data_ref = Reference(ws3, min_col=5, min_row=1, max_row=len(df) + 1)
            dates_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(df) + 1)
            hist_chart.add_data(data_ref, titles_from_data=True)
            hist_chart.set_categories(dates_ref)
            hist_chart.series[0].graphicalProperties.line.solidFill = BRAND_BLUE
            hist_chart.series[0].graphicalProperties.line.width = 18000
            hist_chart.legend = None
            ws3.add_chart(hist_chart, f"I2")

        for i in range(1, 8):
            ws3.column_dimensions[get_column_letter(i)].width = 18

    elif last_values and last_index:
        # Fallback: use stored close values
        hist_headers = ["Date", "Close"]
        for col, h in enumerate(hist_headers, 1):
            ws3.cell(row=1, column=col, value=h)
        style_header_row(ws3, 1, len(hist_headers))

        for i, (dt, val) in enumerate(zip(last_index, last_values), 2):
            ws3.cell(row=i, column=1, value=dt[:19]).font = normal_font
            ws3.cell(row=i, column=1).border = thin_border
            c = ws3.cell(row=i, column=2, value=round(val, 6))
            c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border

        if len(last_values) > 1:
            hist_chart = LineChart()
            hist_chart.title = f"{symbol} Prices (Training Snapshot)"
            hist_chart.style = 10
            hist_chart.width = 30; hist_chart.height = 15
            data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(last_values) + 1)
            dates_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(last_values) + 1)
            hist_chart.add_data(data_ref, titles_from_data=True)
            hist_chart.set_categories(dates_ref)
            hist_chart.series[0].graphicalProperties.line.solidFill = BRAND_BLUE
            hist_chart.series[0].graphicalProperties.line.width = 18000
            hist_chart.legend = None
            ws3.add_chart(hist_chart, "D2")

        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 18

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 4: Combined Chart (History + Forecast)
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Combined Chart")
    ws4.sheet_properties.tabColor = "10B981"

    # Build combined data: last N historical + forecast
    hist_tail_n = min(60, len(last_values)) if last_values else 0
    hist_dates = last_index[-hist_tail_n:] if last_index else []
    hist_prices = last_values[-hist_tail_n:] if last_values else []

    combined_headers = ["Date", "Historical Price", "Forecast Price"]
    for col, h in enumerate(combined_headers, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header_row(ws4, 1, 3)

    chart_row = 2
    for dt, price in zip(hist_dates, hist_prices):
        ws4.cell(row=chart_row, column=1, value=dt[:10]).font = normal_font
        ws4.cell(row=chart_row, column=1).border = thin_border
        c = ws4.cell(row=chart_row, column=2, value=round(price, 6))
        c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
        ws4.cell(row=chart_row, column=3).border = thin_border  # empty forecast column
        chart_row += 1

    # Add a bridge point: last historical = first forecast
    if hist_prices and forecast_points:
        ws4.cell(row=chart_row, column=1, value=hist_dates[-1][:10] if hist_dates else "").font = normal_font
        ws4.cell(row=chart_row, column=1).border = thin_border
        c = ws4.cell(row=chart_row, column=2, value=round(hist_prices[-1], 6))
        c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
        c = ws4.cell(row=chart_row, column=3, value=round(hist_prices[-1], 6))
        c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
        c.fill = forecast_fill
        chart_row += 1

    for fp in forecast_points:
        ws4.cell(row=chart_row, column=1, value=fp.time[:10]).font = normal_font
        ws4.cell(row=chart_row, column=1).border = thin_border
        ws4.cell(row=chart_row, column=2).border = thin_border  # empty historical column
        c = ws4.cell(row=chart_row, column=3, value=round(fp.predicted, 6))
        c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
        c.fill = forecast_fill
        chart_row += 1

    total_rows = chart_row - 1
    if total_rows > 2:
        combo_chart = LineChart()
        combo_chart.title = f"{symbol} — Historical + {model_type} Forecast"
        combo_chart.style = 10
        combo_chart.y_axis.title = "Price"
        combo_chart.width = 35; combo_chart.height = 18

        hist_ref = Reference(ws4, min_col=2, min_row=1, max_row=total_rows + 1)
        fc_ref = Reference(ws4, min_col=3, min_row=1, max_row=total_rows + 1)
        dates_ref = Reference(ws4, min_col=1, min_row=2, max_row=total_rows + 1)

        combo_chart.add_data(hist_ref, titles_from_data=True)
        combo_chart.add_data(fc_ref, titles_from_data=True)
        combo_chart.set_categories(dates_ref)

        combo_chart.series[0].graphicalProperties.line.solidFill = BRAND_BLUE
        combo_chart.series[0].graphicalProperties.line.width = 22000
        combo_chart.series[1].graphicalProperties.line.solidFill = FORECAST_COLOR
        combo_chart.series[1].graphicalProperties.line.width = 22000
        combo_chart.series[1].graphicalProperties.line.dashStyle = "dash"

        ws4.add_chart(combo_chart, "E2")

    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 20

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 5: How to Verify (check tomorrow)
    # ═══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("How to Verify")
    ws5.sheet_properties.tabColor = "EF4444"

    ws5.merge_cells("A1:E1")
    ws5["A1"] = "How to Verify This Forecast"
    ws5["A1"].font = brand_font
    ws5["A1"].alignment = Alignment(horizontal="center")
    ws5.row_dimensions[1].height = 35

    row = 3
    instructions = [
        "This report contains a forecast generated by a machine learning / statistical model.",
        "",
        f"Model: {model_type}",
        f"Currency Pair: {symbol}",
        f"Target: {'Close Price' if target_type == 'close' else 'Log Return (converted to price)'}",
        f"Forecast starts: {forecast_points[0].time if forecast_points else 'N/A'}",
        "",
        "— HOW TO CHECK IF THE FORECAST WAS CORRECT —",
        "",
        "1. Open the 'Forecast Data' sheet in this workbook.",
        "2. For each forecasted date, compare the 'Predicted Price' column",
        "   with the actual closing price you can find on:",
        "   • Google Finance: search '{} to {}'".format(symbol[:3], symbol[3:]),
        "   • Yahoo Finance: search '{}={}'".format(symbol[:3], symbol[3:]),
        "   • TradingView: search '{}/{}'".format(symbol[:3], symbol[3:]),
        "",
        "3. Calculate the error: |Actual - Predicted|",
        "   If the error is less than the MAE shown in the model details,",
        "   the model performed better than its training average.",
        "",
        "4. Check the direction: did the price go UP or DOWN as predicted?",
        "   Directional accuracy is often more useful than exact price match.",
        "",
        "— IMPORTANT DISCLAIMER —",
        "",
        "This forecast is generated by a statistical/ML model for educational",
        "and research purposes only. Currency markets are influenced by many",
        "unpredictable factors. DO NOT use this as financial advice.",
        f"",
        f"Generated: {now.strftime('%Y-%m-%d %H:%M UTC')}",
    ]

    for text in instructions:
        ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws5.cell(row=row, column=1, value=text)
        if text.startswith("—"):
            c.font = subtitle_font
        elif text.startswith("Model:") or text.startswith("Currency") or text.startswith("Target:") or text.startswith("Forecast starts"):
            c.font = Font(name="Calibri", bold=True, size=11)
        else:
            c.font = normal_font
        c.alignment = wrap_align
        row += 1

    # Quick comparison table
    row += 1
    ws5.cell(row=row, column=1, value="Quick Comparison Table").font = title_font
    row += 1

    comp_headers = ["Date", "Predicted Price", "Actual Price (fill in)", "Error", "Direction Correct?"]
    for col, h in enumerate(comp_headers, 1):
        ws5.cell(row=row, column=col, value=h)
    style_header_row(ws5, row, len(comp_headers))
    row += 1

    # First 10 forecast points for quick checking
    for fp in forecast_points[:10]:
        ws5.cell(row=row, column=1, value=fp.time).font = normal_font
        ws5.cell(row=row, column=1).border = thin_border
        c = ws5.cell(row=row, column=2, value=round(fp.predicted, 6))
        c.font = mono_font; c.number_format = "0.000000"; c.border = thin_border
        # Actual price — empty for user to fill
        ws5.cell(row=row, column=3).border = thin_border
        ws5.cell(row=row, column=3).fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        # Error formula
        ws5.cell(row=row, column=4, value=f"=IF(C{row}<>\"\",ABS(C{row}-B{row}),\"\")").font = mono_font
        ws5.cell(row=row, column=4).border = thin_border
        ws5.cell(row=row, column=4).number_format = "0.000000"
        # Direction
        ws5.cell(row=row, column=5).border = thin_border
        row += 1

    ws5.column_dimensions["A"].width = 15
    ws5.column_dimensions["B"].width = 20
    ws5.column_dimensions["C"].width = 22
    ws5.column_dimensions["D"].width = 15
    ws5.column_dimensions["E"].width = 22

    # ═══════════════════════════════════════════════════════════════════════════
    # Save
    # ═══════════════════════════════════════════════════════════════════════════
    report_path = settings.reports_path / f"forecast_{report_id}.xlsx"
    wb.save(str(report_path))

    filename = f"forecast_{symbol}_{model_type}_{now.strftime('%Y%m%d')}.xlsx"
    log.info("forecast_excel.generated", report_id=report_id, path=str(report_path))
    return report_id, report_path, filename
